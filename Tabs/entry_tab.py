import os
import datetime
import pandas as pd
import customtkinter as ctk
from tkinter import messagebox
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import calendar


def create_monthly_excel_template(customers_df, year, month, path, month_rate=0):
    days_in_month = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}_{month:02d}"

    ws['A1'] = 'S.No'
    ws['B1'] = 'CID'
    ws['C1'] = 'Name'
    ws['D1'] = 'Phone'
    ws['E1'] = month_rate
    ws['E1'].font = Font(bold=True)

    col_start = 6
    col_end = col_start + days_in_month - 1

    for col in range(col_start, col_end + 1):
        day_label = datetime.date(year, month, col - col_start + 1).strftime('%d/%m/%y')
        cell = ws.cell(row=1, column=col, value=day_label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 8

    total_ltr_col = col_end + 1
    total_rt_col = total_ltr_col + 1

    ws.cell(row=1, column=total_ltr_col, value='Total_LTR').font = Font(bold=True)
    ws.cell(row=1, column=total_rt_col, value='Total_RT').font = Font(bold=True)
    ws.column_dimensions[get_column_letter(total_ltr_col)].width = 15
    ws.column_dimensions[get_column_letter(total_rt_col)].width = 15

    max_customers = max(500, len(customers_df))
    start_row = 2

    for i in range(max_customers):
        row_num = start_row + i
        if i < len(customers_df):
            cust = customers_df.iloc[i]
            ws.cell(row=row_num, column=1, value=i + 1)
            ws.cell(row=row_num, column=2, value=cust['CID'])
            ws.cell(row=row_num, column=3, value=cust['Name'])
            ws.cell(row=row_num, column=4, value=cust.get('Phone', None))
        else:
            ws.cell(row=row_num, column=1, value=i + 1)

        for col in range(col_start, col_end + 1):
            ws.cell(row=row_num, column=col, value=None)

        first_col_letter = get_column_letter(col_start)
        last_col_letter = get_column_letter(col_end)
        total_ltr_col_letter = get_column_letter(total_ltr_col)

        ws.cell(row=row_num, column=total_ltr_col).value = f"=SUM({first_col_letter}{row_num}:{last_col_letter}{row_num})"
        ws.cell(row=row_num, column=total_rt_col).value = f"=IF(ISNUMBER({total_ltr_col_letter}{row_num}), {total_ltr_col_letter}{row_num}*$E$1, 0)"

    wb.save(path)

def reload_customers(self):
    self.load_customers()

def _begin_page_build(self):
    self.pack_forget()
    self.update_idletasks()

def _end_page_build(self):
    self.pack(fill="both", expand=True)
    self.update_idletasks()
    




def sync_customers_to_monthly(customers_path, monthly_path, year, month, month_rate=0):
    """Sync customers - NEVER add new customers to old sheets"""
    
    main_customers = pd.read_excel(customers_path)[['CID', 'Name', 'Phone']]

    if not os.path.exists(monthly_path):
        create_monthly_excel_template(main_customers, year, month, monthly_path, month_rate)
        return

    try:
        wb = load_workbook(monthly_path)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Load Error", f"Could not load file: {str(e)}")
        return

    days_in_month = calendar.monthrange(year, month)[1]
    col_start = 6
    col_end = col_start + days_in_month - 1
    total_ltr_col = col_end + 1
    total_rt_col = total_ltr_col + 1

    existing_cids = {}
    for row in range(2, ws.max_row + 1):
        cid_cell = ws.cell(row=row, column=2).value
        if cid_cell:
            existing_cids[cid_cell] = row

    # ✅ CHECK IF THIS IS AN OLD SHEET
    today = datetime.date.today()
    sheet_date = datetime.date(year, month, 1)
    is_old_sheet = sheet_date < datetime.date(today.year, today.month, 1)

    # ✅ ONLY ADD NEW CUSTOMERS TO CURRENT/FUTURE SHEETS
    if not is_old_sheet:
        new_customers = main_customers[~main_customers['CID'].isin(existing_cids.keys())]

        if not new_customers.empty:
            last_filled_row = max(existing_cids.values()) if existing_cids else 1

            for idx, (_, row_data) in enumerate(new_customers.iterrows()):
                new_row = last_filled_row + idx + 1

                ws.cell(row=new_row, column=1, value=new_row - 1)
                ws.cell(row=new_row, column=2, value=row_data['CID'])
                ws.cell(row=new_row, column=3, value=row_data['Name'])
                ws.cell(row=new_row, column=4, value=row_data.get('Phone', None))

                for col in range(col_start, col_end + 1):
                    ws.cell(row=new_row, column=col, value=None)

                first_col_letter = get_column_letter(col_start)
                last_col_letter = get_column_letter(col_end)
                total_ltr_col_letter = get_column_letter(total_ltr_col)

                ws.cell(row=new_row, column=total_ltr_col).value = f"=SUM({first_col_letter}{new_row}:{last_col_letter}{new_row})"
                ws.cell(row=new_row, column=total_rt_col).value = f"=IF(ISNUMBER({total_ltr_col_letter}{new_row}), {total_ltr_col_letter}{new_row}*$E$1, 0)"

    # ALWAYS update name/phone for existing customers (even in old sheets)
    cid_to_name = pd.Series(main_customers['Name'].values, index=main_customers['CID']).to_dict()
    cid_to_phone = pd.Series(main_customers['Phone'].values, index=main_customers['CID']).to_dict()

    for row in range(2, ws.max_row + 1):
        cid_cell = ws.cell(row=row, column=2).value
        if cid_cell:
            ws.cell(row=row, column=1, value=row - 1)
            
            name_value = cid_to_name.get(cid_cell, None)
            if name_value:
                ws.cell(row=row, column=3, value=name_value)
            
            phone_value = cid_to_phone.get(cid_cell, None)
            ws.cell(row=row, column=4, value=phone_value)

    # Ensure formulas
    for row in range(2, ws.max_row + 1):
        cid_cell = ws.cell(row=row, column=2).value
        if cid_cell:
            first_col_letter = get_column_letter(col_start)
            last_col_letter = get_column_letter(col_end)
            total_ltr_col_letter = get_column_letter(total_ltr_col)

            current_formula = ws.cell(row=row, column=total_ltr_col).value
            if not current_formula or not str(current_formula).startswith('='):
                ws.cell(row=row, column=total_ltr_col).value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"

            current_rt_formula = ws.cell(row=row, column=total_rt_col).value
            if not current_rt_formula or not str(current_rt_formula).startswith('='):
                ws.cell(row=row, column=total_rt_col).value = f"=IF(ISNUMBER({total_ltr_col_letter}{row}), {total_ltr_col_letter}{row}*$E$1, 0)"

    try:
        wb.save(monthly_path)
    except PermissionError:
        messagebox.showerror("Save Error", f"Close the file {monthly_path} before syncing.")
        return


class EntryTab(ctk.CTkFrame):
    def __init__(self, parent, colors, back_callback, customer_tab):
        
        super().__init__(parent)
        self.parent = parent
        self.colors = colors
        self.back_callback = back_callback
        self.customer_tab = customer_tab

        # UPDATED: Use dynamic paths from app_config
        from app_config import app_config
        self.app_config = app_config
        
        self.monthly_sheets_path = str(app_config.monthly_sheets_path)
        
        # Ensure monthly sheets directory exists
        app_config.monthly_sheets_path.mkdir(parents=True, exist_ok=True)

        self.current_year = datetime.date.today().year
        self.month_names = [
            "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
            "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"
        ]
        self.pack_forget()
        self.init_ui()
    
    def show_save_reminder_toast(self):
        """Show save reminder toast once per session (auto closes)"""

        # Show only once per session
        if hasattr(self, "_save_reminder_shown") and self._save_reminder_shown:
            return
        self._save_reminder_shown = True

        toast = ctk.CTkToplevel(self)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)

        width, height = 450, 150
        x = toast.winfo_screenwidth() - width - 40
        y = toast.winfo_screenheight() - height - 80
        toast.geometry(f"{width}x{height}+{x}+{y}")

        frame = ctk.CTkFrame(
            toast,
            fg_color="#1A2711",
            corner_radius=16,
            border_width=2,
            border_color="#f59e0b"
        )
        frame.pack(fill="both", expand=True)

        ctk.CTkLabel(
            frame,
            text="⚠ IMPORTANT",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#fbbf24"
        ).pack(anchor="w", padx=18, pady=(14, 4))

        ctk.CTkLabel(
            frame,
            text="Always SAVE the Excel sheet after entering data.\n"
                "Unsaved sheets may cause reports to show incorrect values.",
            font=ctk.CTkFont(size=14),
            text_color="#e5e7eb",
            justify="left",
            wraplength=380
        ).pack(anchor="w", padx=18)

        toast.after(5000, toast.destroy)


    def init_ui(self):
        label = ctk.CTkLabel(self, text="Select Year:", font=ctk.CTkFont(size=20, weight="bold"))
        label.pack(pady=(15, 5))

        # Year range 2020 -> 2050
        year_values = [str(y) for y in range(2020, 2051)]
        default_year = str(self.current_year) if str(self.current_year) in year_values else year_values[-1]

        self.year_var = ctk.StringVar(value=default_year)
        year_selector = ctk.CTkComboBox(
            self, 
            values=year_values, 
            variable=self.year_var, 
            width=150, 
            height=36,
            font=ctk.CTkFont(size=14)
        )
        year_selector.pack()

        label2 = ctk.CTkLabel(
            self, 
            text="Select Month to Open Excel Sheet", 
            font=ctk.CTkFont(size=24, weight="bold")
        )
        label2.pack(pady=20)

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack()

        current_month = datetime.date.today().month

        for i, month_name in enumerate(self.month_names, 1):

            if i == current_month:
                fg = "#22c55e"          # green highlight
                hover = "#16a34a"
                border_color = "#14532d"
                border_width = 3
            else:
                fg = self.colors["primary"]
                hover = "#1e40af"
                border_color = "#cbd5f5"
                border_width = 1

            button = ctk.CTkButton(
                btn_frame,
                text=month_name,
                width=140,
                height=45,
                font=ctk.CTkFont(size=14, weight="bold"),
                corner_radius=14,
                fg_color=fg,
                hover_color=hover,
                border_color=border_color,
                border_width=border_width,
                text_color="#ffffff",
                command=lambda m=i: self._on_month_clicked(m)
            )

            button.grid(
                row=(i - 1) // 4,
                column=(i - 1) % 4,
                padx=10,
                pady=10
            )



        back_button = ctk.CTkButton(
            self, 
            text="Back", 
            width=120, 
            height=40, 
            command=self.back_callback
        )
        back_button.pack(pady=25)

    def _on_month_clicked(self, month_index):
        """Handle month click: show toast + open sheet"""
        self.show_save_reminder_toast()
        self.open_month_excel(month_index)


    def open_month_excel(self, month):
        year = int(self.year_var.get())
        monthly_path = os.path.join(self.monthly_sheets_path, f"{year}_{month:02d}.xlsx")

        if not os.path.exists(monthly_path):
            create_new = messagebox.askyesno(
                "No Data", 
                f"No monthly sheet exists for {self.month_names[month-1]} {year}. Create new sheet?"
            )
            if not create_new:
                return
            
            customers_path = self.customer_tab.excel_file
            
            # Check if customers file exists
            if not os.path.exists(customers_path):
                messagebox.showerror(
                    "Error",
                    f"Customers file not found at:\n{customers_path}\n\nPlease add customers first.",
                    parent=self
                )
                return
            
            try:
                main_customers = pd.read_excel(customers_path)[['CID', 'Name', 'Phone']]
                create_monthly_excel_template(main_customers, year, month, monthly_path)
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Failed to create monthly sheet:\n{str(e)}",
                    parent=self
                )
                return
        else:
            customers_path = self.customer_tab.excel_file

        sync_customers_to_monthly(customers_path, monthly_path, year, month)

        try:
            if os.name == 'nt':
                os.startfile(monthly_path)
            else:
                subprocess.call(['open', monthly_path])
        except Exception as e:
            messagebox.showerror("Open File Error", f"Could not open file:\n{str(e)}")