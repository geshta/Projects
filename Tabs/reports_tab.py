import os
import datetime
import pandas as pd
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import urllib.parse


class ReportTab(ctk.CTkFrame):
    def __init__(self, parent, colors, back_callback):
        
        super().__init__(parent)
        self.parent = parent
        self.colors = colors
        self.back_callback = back_callback
        
        # UPDATED: Use dynamic paths from app_config
        from app_config import app_config
        self.app_config = app_config
        
        self.monthly_sheets_path = str(app_config.monthly_sheets_path)
        self.customers_file = str(app_config.customers_file)
        
        self.current_year = datetime.date.today().year
        self.month_names = [
            "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
            "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"
        ]
        
        self.all_customers = []
        self.report_data = []
        self.selected_month = None
        self.selected_year = None
        self.view_mode = "menu"
        
        # UPDATED: Get business info from config
        self.business_name = app_config.get_business_name()
        self.contact_number = app_config.get_contact_number()
        self.payment_info = app_config.get_payment_info()
        self.pack_forget()    
        self.init_menu_ui()

    def _begin_page_build(self):
        self.pack_forget()
        self.update_idletasks()

    def _end_page_build(self):
        self.pack(fill="both", expand=True)
        self.update_idletasks()

    
    def init_menu_ui(self):
        """Show main menu with options"""

        for widget in self.winfo_children():
            widget.destroy()

        self.view_mode = "menu"

        # ================= HEADER =================
        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkButton(
            header,
            text="Back",
            font=ctk.CTkFont(size=24, weight="bold"),
            width=150,
            height=55,
            corner_radius=15,
            fg_color=self.colors["text_light"],
            text_color=self.colors["primary"],
            hover_color="#e2e8f0",
            command=self.back_callback,
        ).pack(side="left", padx=30, pady=12)

        ctk.CTkLabel(
            header,
            text="📊 Customer Reports & Downloads",
            font=ctk.CTkFont(size=40, weight="bold"),
            text_color=self.colors["text_light"],
        ).pack(side="left", padx=20)

        # ================= CONTENT =================
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=40, pady=40)

        ctk.CTkLabel(
            content,
            text="Select Report Type",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=self.colors["text_dark"]
        ).pack(pady=30)

        # ================= CARD GRID =================
        grid = ctk.CTkFrame(content, fg_color="transparent")
        grid.pack(fill="both", expand=True)

        grid.grid_columnconfigure((0, 1, 2), weight=1)

        CARD_W, CARD_H = 330, 230
        CARD_RADIUS = 22

        # ========= CARD CREATION INLINE =========
        def make_card(col, accent, icon, title, desc, command):
            card = ctk.CTkFrame(
                grid,
                fg_color="#ffffff",
                corner_radius=CARD_RADIUS,
                border_width=2,
                border_color="#e5e7eb",
                width=CARD_W,
                height=CARD_H
            )
            card.grid(row=0, column=col, padx=18, pady=10, sticky="nsew")
            card.pack_propagate(False)

            accent_bar = ctk.CTkFrame(card, fg_color=accent, width=6)
            accent_bar.pack(side="left", fill="y")

            body = ctk.CTkFrame(card, fg_color="transparent")
            body.pack(fill="both", expand=True, padx=20, pady=20)

            icon_lbl = ctk.CTkLabel(body, text=icon, font=ctk.CTkFont(size=44), text_color=accent)
            icon_lbl.pack(anchor="w")

            title_lbl = ctk.CTkLabel(body, text=title, font=ctk.CTkFont(size=22, weight="bold"))
            title_lbl.pack(anchor="w", pady=(10, 4))

            desc_lbl = ctk.CTkLabel(
                body,
                text=desc,
                font=ctk.CTkFont(size=14),
                text_color="#6b7280",
                wraplength=260,
                justify="left"
            )
            desc_lbl.pack(anchor="w")

            # ---------- HOVER + CLICK + KEYBOARD ----------
            widgets = [card, body, icon_lbl, title_lbl, desc_lbl, accent_bar]

            def on_enter(e=None):
                card.configure(border_color=accent)

            def on_leave(e=None):
                card.configure(border_color="#e5e7eb")

            def on_click(e=None):
                command()

            for w in widgets:
                w.bind("<Enter>", on_enter)
                w.bind("<Leave>", on_leave)
                w.bind("<Button-1>", on_click)
                w.bind("<Return>", on_click)
                w.configure(cursor="hand2")

            card.focus_set()

        # -------- CREATE CARDS --------
        make_card(
            0,
            "#10b981",
            "📋",
            "All Records",
            "View every customer with full lifetime totals and download options.",
            self.show_all_records_view
        )

        make_card(
            1,
            "#3b82f6",
            "📅",
            "Monthly Report",
            "Generate detailed milk delivery and billing report for a selected month.",
            self.show_month_selection
        )

        make_card(
            2,
            "#f59e0b",
            "📆",
            "Yearly Report",
            "See year-wise summary of customer totals across all months.",
            self.show_year_selection_for_yearly
        )




   
    def _focus_first_row(self):
        try:
            self.tree.focus_set()
            items = self.tree.get_children()
            if items:
                self.tree.focus(items[0])
                self.tree.selection_set(items[0])
                self.tree.see(items[0])
        except:
            pass
    
    def load_customers(self):
        try:
            self.tree.delete(*self.tree.get_children())

            df = pd.read_excel(self.excel_file, engine="openpyxl")
            if df.empty:
                return

            df = df.astype(str)

            for idx, row in enumerate(df.itertuples(index=False), start=1):
                tag = "oddrow" if idx % 2 else "evenrow"
                self.tree.insert(
                    "",
                    "end",
                    values=(idx, row.CID, row.Name, row.Phone, row.Address),
                    tags=(tag,)
                )

            # Focus first row for keyboard navigation
            items = self.tree.get_children()
            if items:
                self.tree.focus_set()
                self.tree.focus(items[0])
                self.tree.selection_set(items[0])
                self.tree.see(items[0])

        except Exception as e:
            messagebox.showerror("Load Error", str(e), parent=self)




    def show_month_selection(self):
        """Show month selection screen"""
        for widget in self.winfo_children():
            widget.destroy()
        
        # Header
        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        btn_back = ctk.CTkButton(
            header,
            text="Back",
            font=ctk.CTkFont(size=24, weight="bold"),
            width=150,
            height=55,
            corner_radius=15,
            fg_color=self.colors["text_light"],
            text_color=self.colors["primary"],
            hover_color="#e2e8f0",
            command=self.init_menu_ui,
        )
        btn_back.pack(side="left", padx=30, pady=12)
        
        lbl_title = ctk.CTkLabel(
            header,
            text="📅 Select Month for Report",
            font=ctk.CTkFont(size=36, weight="bold"),
            text_color=self.colors["text_light"],
        )
        lbl_title.pack(side="left", padx=20)
        
        # Content
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=40, pady=20)
        
        # Year selector
        year_frame = ctk.CTkFrame(content, fg_color="transparent")
        year_frame.pack(pady=(10, 5))
        
        ctk.CTkLabel(
            year_frame,
            text="Select Year:",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=self.colors["text_dark"]
        ).pack(side="left", padx=10)
        
        year_values = [str(y) for y in range(2020, 2051)]
        self.year_var = ctk.StringVar(value=str(self.current_year))
        year_selector = ctk.CTkComboBox(
            year_frame,
            values=year_values,
            variable=self.year_var,
            width=150,
            height=40,
            corner_radius=10,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        year_selector.pack(side="left", padx=10)
        
        # Month buttons
        ctk.CTkLabel(
            content,
            text="Select Month",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=self.colors["text_dark"]
        ).pack(pady=20)
        
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack()
        
        current_month = datetime.date.today().month

        for i, month_name in enumerate(self.month_names, 1):

            if i == current_month:
                fg = "#22c55e"
                hover = "#16a34a"
                border_color = "#14532d"
                border_width = 3
            else:
                fg = self.colors["primary"]
                hover = "#1e40af"
                border_color = "#cbd5f5"
                border_width = 1

            button = ctk.CTkButton(
                btn_frame,
                text=month_name,
                width=140,
                height=45,
                font=ctk.CTkFont(size=14, weight="bold"),
                corner_radius=14,
                fg_color=fg,
                hover_color=hover,
                border_color=border_color,
                border_width=border_width,
                text_color="#ffffff",
                command=lambda m=i: self.show_monthly_data(m)
            )

            button.grid(
                row=(i - 1) // 4,
                column=(i - 1) % 4,
                padx=10,
                pady=10
            )


    def get_all_customers_from_history(self):
        """Get ALL customers including deleted ones by scanning monthly sheets
        UPDATED: Always get latest name/phone/address from main customers file"""
        all_unique_customers = {}
        
        # First, add current customers from main sheet
        try:
            df_customers = pd.read_excel(self.customers_file, engine="openpyxl")
            for _, row in df_customers.iterrows():
                cid = str(row.get('CID', '')).strip()
                if cid:
                    all_unique_customers[cid] = {
                        'CID': cid,
                        'Name': row.get('Name', ''),
                        'Phone': row.get('Phone', ''),
                        'Address': row.get('Address', ''),
                        'Status': 'Active'
                    }
        except Exception as e:
            print(f"Error loading customers: {e}")
        
        # Now scan ALL monthly sheets to find deleted customers
        if os.path.exists(self.monthly_sheets_path):
            for filename in os.listdir(self.monthly_sheets_path):
                if filename.endswith('.xlsx'):
                    filepath = os.path.join(self.monthly_sheets_path, filename)
                    try:
                        wb = load_workbook(filepath, data_only=True)
                        ws = wb.active
                        
                        for row_idx in range(2, ws.max_row + 1):
                            cid = ws.cell(row=row_idx, column=2).value
                            name = ws.cell(row=row_idx, column=3).value
                            phone = ws.cell(row=row_idx, column=4).value
                            
                            if not cid:
                                continue
                            
                            cid_str = str(cid).strip()
                            
                            # If this CID not in our dict, it's a deleted customer
                            if cid_str not in all_unique_customers:
                                all_unique_customers[cid_str] = {
                                    'CID': cid_str,
                                    'Name': name or '',
                                    'Phone': phone or '',
                                    'Address': '',
                                    'Status': 'Deleted (Has History)'
                                }
                        
                        wb.close()
                    except Exception as e:
                        print(f"Error reading {filename}: {e}")
                        continue
        
        return list(all_unique_customers.values())


    def show_all_records_view(self):
        """Display all customers with cumulative totals"""
        for widget in self.winfo_children():
            widget.destroy()
        
        self.view_mode = "all_records"
        
        # Load ALL customers including deleted ones
        try:
            self.all_customers = self.get_all_customers_from_history()
            self.calculate_total_liters()
            self.debug_check_data()
            self.all_customers = sorted(self.all_customers, key=lambda x: str(x.get("CID", "")).zfill(10))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load customers: {e}", parent=self)
            self.init_menu_ui()
            return
        
        # ==================== HEADER ====================
        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        btn_back = ctk.CTkButton(
            header,
            text="Back",
            font=ctk.CTkFont(size=24, weight="bold"),
            width=150,
            height=55,
            corner_radius=15,
            fg_color=self.colors["text_light"],
            text_color=self.colors["primary"],
            hover_color="#e2e8f0",
            command=self.init_menu_ui,
        )
        btn_back.pack(side="left", padx=30, pady=12)
        
        lbl_title = ctk.CTkLabel(
            header,
            text="📋 All Customer Records",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=self.colors["text_light"],
        )
        lbl_title.pack(side="left", padx=20)
        
        # ==================== TOOLBAR ====================
        toolbar = ctk.CTkFrame(self, fg_color=self.colors["secondary"], height=120)
        toolbar.pack(fill="x", padx=20, pady=(10, 5))
        toolbar.pack_propagate(False)

        # -------- LEFT: Search Frame --------
        search_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        search_frame.pack(side="left", padx=20, pady=10)

        self.ent_search = ctk.CTkEntry(
            search_frame,
            placeholder_text="Search CID, Name, Phone, Address",
            font=ctk.CTkFont(size=16),
            width=400,
            height=45,
            corner_radius=15,
            border_width=2,
            border_color=self.colors["info"],
            fg_color="#ffffff",
            text_color="#000000",
        )
        self.ent_search.pack(side="left", padx=5)

        btn_search = ctk.CTkButton(
            search_frame,
            text="Search",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color=self.colors["info"],
            hover_color="#0891b2",
            command=self.search_customers,
        )
        btn_search.pack(side="left", padx=5)

        btn_clear = ctk.CTkButton(
            search_frame,
            text="Clear",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=self.clear_search,
        )
        btn_clear.pack(side="left", padx=5)

        btn_refresh = ctk.CTkButton(
            search_frame,
            text="Refresh",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=120,
            height=45,
            corner_radius=12,
            fg_color="#f59e0b",
            hover_color="#d97706",
            command=self.refresh_all_records
        )
        btn_refresh.pack(side="left", padx=5)

        # -------- RIGHT: Stats + Download --------
        right_container = ctk.CTkFrame(toolbar, fg_color="transparent")
        right_container.pack(side="right", padx=20, pady=10)

        # Calculate totals
        total_liters = sum(float(c.get('Total_Liters', 0)) for c in self.all_customers)
        total_amount = sum(float(c.get('Total_Amount', 0)) for c in self.all_customers)

        # ✅ TOTALS BOX - FIXED HEIGHT + NO PROPAGATE
        totals_frame = ctk.CTkFrame(
            right_container,
            fg_color="#1e293b",
            corner_radius=15,
            width=220,
            height=100
        )
        totals_frame.pack(side="left", padx=(0, 12))
        totals_frame.pack_propagate(False)  # ← CRITICAL

        ctk.CTkLabel(
            totals_frame,
            text="📊 CUMULATIVE TOTALS",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#fbbf24"
        ).pack(pady=(10, 5))

        ctk.CTkLabel(
            totals_frame,
            text=f"🥛 {total_liters:.2f} L",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#10b981"
        ).pack(pady=2)

        ctk.CTkLabel(
            totals_frame,
            text=f"💰 ₹{total_amount:.2f}",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#3b82f6"
        ).pack(pady=(2, 10))

        # ✅ DOWNLOAD BUTTON - MATCHING HEIGHT
        btn_download = ctk.CTkButton(
            right_container,
            text="📥 DOWNLOAD",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=180,
            height=100,
            corner_radius=15,
            fg_color="#10b981",
            hover_color="#059669",
            command=self.download_all_records
        )
        btn_download.pack(side="left")
        
        # ==================== TABLE ====================
        self.create_report_table()
        self.report_data = self.all_customers.copy()
        self.populate_table(self.report_data)
        
        # ==================== BOTTOM STATUS ====================
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent", height=70)
        bottom_frame.pack(fill="x", padx=20, pady=10)
        bottom_frame.pack_propagate(False)
        
        active_count = sum(1 for c in self.all_customers if c.get('Status') == 'Active')
        deleted_count = len(self.all_customers) - active_count
        
        self.lbl_status = ctk.CTkLabel(
            bottom_frame,
            text=f"Showing {len(self.all_customers)} customers (Active: {active_count}, Deleted: {deleted_count})",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=self.colors["text_dark"]
        )
        self.lbl_status.pack(side="left", padx=20)






    def show_monthly_data(self, month):
        """Display monthly data for selected month
        UPDATED: Always fetch latest name/phone/address from main customers file"""
        for widget in self.winfo_children():
            widget.destroy()
        
        self.selected_month = month
        try:
            self.selected_year = int(self.year_var.get())
        except:
            self.selected_year = self.current_year
        
        monthly_file = os.path.join(
            self.monthly_sheets_path,
            f"{self.selected_year}_{self.selected_month:02d}.xlsx"
        )
        
        if not os.path.exists(monthly_file):
            messagebox.showwarning(
                "File Not Found",
                f"❌ No data found for {self.month_names[month-1]} {self.selected_year}",
                parent=self
            )
            self.show_month_selection()
            return
        
        # Load monthly data
        try:
            # FIRST: Load latest customer data from main file
            customers_dict = {}
            try:
                df_customers = pd.read_excel(self.customers_file, engine="openpyxl")
                for _, row in df_customers.iterrows():
                    cid = str(row.get('CID', '')).strip()
                    if cid:
                        customers_dict[cid] = {
                            'Name': row.get('Name', ''),
                            'Phone': row.get('Phone', ''),
                            'Address': row.get('Address', '')
                        }
            except Exception as e:
                print(f"Error loading main customers: {e}")
            
            # THEN: Load monthly sheet
            wb = load_workbook(monthly_file, data_only=True)
            ws = wb.active
            
            days_in_month = calendar.monthrange(self.selected_year, self.selected_month)[1]
            col_start = 6
            total_ltr_col = col_start + days_in_month
            total_amt_col = total_ltr_col + 1
            
            self.report_data = []
            
            for row_idx in range(2, ws.max_row + 1):
                cid = ws.cell(row=row_idx, column=2).value
                name_sheet = ws.cell(row=row_idx, column=3).value
                
                if not cid or not name_sheet:
                    continue
                
                cid_str = str(cid).strip()
                
                ltr_val = ws.cell(row=row_idx, column=total_ltr_col).value
                amt_val = ws.cell(row=row_idx, column=total_amt_col).value

                # ---------- FIX: handle Excel not saved ----------
                total_ltr = 0.0
                total_amt = 0.0

                try:
                    if ltr_val is not None and amt_val is not None:
                        total_ltr = float(ltr_val)
                        total_amt = float(amt_val)
                    else:
                        # Recalculate liters manually from daily columns
                        for dcol in range(6, total_ltr_col):
                            day_val = ws.cell(row=row_idx, column=dcol).value
                            if day_val:
                                total_ltr += float(day_val)

                        # ---------- FIX: get rate from customers master ----------
                        rate = 0.0
                        try:
                            # match customer from loaded master data
                            for cust in all_customers:
                                if str(cust.get("CID", "")).strip() == str(cid).strip():
                                    rate = float(cust.get("Rate", 0))
                                    break
                        except:
                            rate = 0.0

                        total_amt = total_ltr * rate

                except:
                    total_ltr = 0.0
                    total_amt = 0.0

                
                # Get LATEST data from main customers file
                if cid_str in customers_dict:
                    name = customers_dict[cid_str]['Name']
                    phone = customers_dict[cid_str]['Phone']
                    address = customers_dict[cid_str]['Address']
                else:
                    # Fallback to sheet data for deleted customers
                    name = name_sheet
                    phone = ws.cell(row=row_idx, column=4).value or ''
                    address = ''
                
                self.report_data.append({
                    'CID': cid,
                    'Name': name,
                    'Phone': phone,
                    'Address': address,
                    'Total_Liters': total_ltr,
                    'Total_Amount': total_amt,
                })
            
            wb.close()
            self.report_data = sorted(self.report_data, key=lambda x: str(x.get("CID", "")).zfill(10))
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load monthly data: {e}", parent=self)
            self.show_month_selection()
            return
        
        # Build UI (rest of the function remains the same)
        self.view_mode = "monthly"
        
        # Header
        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        btn_back = ctk.CTkButton(
            header,
            text="Back",
            font=ctk.CTkFont(size=24, weight="bold"),
            width=150,
            height=55,
            corner_radius=15,
            fg_color=self.colors["text_light"],
            text_color=self.colors["primary"],
            hover_color="#e2e8f0",
            command=self.show_month_selection,
        )
        btn_back.pack(side="left", padx=30, pady=12)
        
        lbl_title = ctk.CTkLabel(
            header,
            text=f"📅 {self.month_names[month-1]} {self.selected_year} Report",
            font=ctk.CTkFont(size=36, weight="bold"),
            text_color=self.colors["text_light"],
        )
        lbl_title.pack(side="left", padx=20)
        
        # Toolbar
        toolbar = ctk.CTkFrame(self, fg_color=self.colors["secondary"], height=120)
        toolbar.pack(fill="x", padx=20, pady=(10, 5))
        toolbar.pack_propagate(False)
        
        # Search
        search_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        search_frame.pack(side="left", padx=20, pady=10)
        
        self.ent_search = ctk.CTkEntry(
            search_frame,
            placeholder_text="Search CID, Name, Phone",
            font=ctk.CTkFont(size=16),
            width=350,
            height=45,
            corner_radius=15,
            border_width=2,
            border_color=self.colors["info"],
            fg_color="#ffffff",
            text_color="#000000",
        )
        self.ent_search.pack(side="left", padx=5)
        
        btn_search = ctk.CTkButton(
            search_frame,
            text="Search",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color=self.colors["info"],
            hover_color="#0891b2",
            command=self.search_customers,
        )
        btn_search.pack(side="left", padx=5)
        
        btn_clear = ctk.CTkButton(
            search_frame,
            text="Clear",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=self.clear_search,
        )
        btn_clear.pack(side="left", padx=5)

        btn_refresh = ctk.CTkButton(
            search_frame,
            text="Refresh",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=120,
            height=45,
            corner_radius=12,
            fg_color="#f59e0b",
            hover_color="#d97706",
            command=lambda: self.show_monthly_data(self.selected_month)
        )
        btn_refresh.pack(side="left", padx=5)
        
        # Download button
        btn_download = ctk.CTkButton(
            toolbar,
            text="📥 DOWNLOAD",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=200,
            height=55,
            corner_radius=14,
            fg_color="#10b981",
            hover_color="#059669",
            command=self.show_download_options
        )
        btn_download.pack(side="right", padx=(10, 20), pady=20)
        
        # Monthly Totals
        totals_frame = ctk.CTkFrame(toolbar, fg_color="#1e293b", corner_radius=15)
        totals_frame.pack(side="right", padx=0, pady=10)
        
        month_total_liters = sum(float(c.get('Total_Liters', 0)) for c in self.report_data)
        month_total_amount = sum(float(c.get('Total_Amount', 0)) for c in self.report_data)
        
        ctk.CTkLabel(
            totals_frame,
            text=f"📊 {self.month_names[month-1].upper()} TOTALS",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#fbbf24"
        ).pack(padx=20, pady=(10, 5))
        
        ctk.CTkLabel(
            totals_frame,
            text=f"🥛 Total Liters: {month_total_liters:.2f} L",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#10b981"
        ).pack(padx=20, pady=2)
        
        ctk.CTkLabel(
            totals_frame,
            text=f"💰 Total Amount: ₹{month_total_amount:.2f}",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#3b82f6"
        ).pack(padx=20, pady=(2, 10))
        
        # Table
        self.create_report_table()
        self.populate_table(self.report_data)
        
        # Bottom frame
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent", height=70)
        bottom_frame.pack(fill="x", padx=20, pady=10)
        bottom_frame.pack_propagate(False)
        
        self.lbl_status = ctk.CTkLabel(
            bottom_frame,
            text=f"Showing {len(self.report_data)} customers for {self.month_names[month-1]} {self.selected_year}",
            font=ctk.CTkFont(size=14),
            text_color=self.colors["text_dark"]
        )
        self.lbl_status.pack(side="left", padx=20)

    def refresh_all_records(self):
        """Reload all customer records and refresh table + totals"""
        try:
            # Reload customers (including deleted)
            self.all_customers = self.get_all_customers_from_history()

            # Recalculate totals from monthly sheets
            self.calculate_total_liters()

            # Sort by CID
            self.all_customers = sorted(
                self.all_customers,
                key=lambda x: str(x.get("CID", "")).zfill(10)
            )

            # Update report data
            self.report_data = self.all_customers.copy()

            # CLEAR table
            for item in self.tree.get_children():
                self.tree.delete(item)

            # REPULATE table
            self.populate_table(self.report_data)

            # Update totals + status
            active_count = sum(1 for c in self.all_customers if c.get('Status') == 'Active')
            deleted_count = len(self.all_customers) - active_count

            total_liters = sum(float(c.get('Total_Liters', 0)) for c in self.all_customers)
            total_amount = sum(float(c.get('Total_Amount', 0)) for c in self.all_customers)

            self.lbl_status.configure(
                text=(
                    f"Showing {len(self.all_customers)} customers "
                    f"(Active: {active_count}, Deleted: {deleted_count}) | "
                    f"🥛 {total_liters:.2f} L | 💰 ₹{total_amount:.2f}"
                )
            )

        except Exception as e:
            messagebox.showerror("Refresh Error", str(e), parent=self)
    

    def _on_arrow_up(self, event):
        items = self.tree.get_children()
        if not items:
            return "break"

        current = self.tree.focus()
        if not current:
            self.tree.focus(items[0])
            self.tree.selection_set(items[0])
            return "break"

        index = items.index(current)
        if index > 0:
            prev_item = items[index - 1]
            self.tree.focus(prev_item)
            self.tree.selection_set(prev_item)
            self.tree.see(prev_item)

        return "break"


    def _on_arrow_down(self, event):
        items = self.tree.get_children()
        if not items:
            return "break"

        current = self.tree.focus()
        if not current:
            self.tree.focus(items[0])
            self.tree.selection_set(items[0])
            return "break"

        index = items.index(current)
        if index < len(items) - 1:
            next_item = items[index + 1]
            self.tree.focus(next_item)
            self.tree.selection_set(next_item)
            self.tree.see(next_item)

        return "break"



    def create_report_table(self):
        table_frame = ctk.CTkFrame(self, fg_color=self.colors["text_light"])
        table_frame.pack(fill="both", expand=True, padx=25, pady=(0, 10))

        table_container = ctk.CTkFrame(table_frame, fg_color="transparent")
        table_container.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Report.Treeview",
            background="#ffffff",
            foreground=self.colors["text_dark"],
            fieldbackground="#ffffff",
            rowheight=50,
            font=("Arial", 14),
        )
        style.configure(
            "Report.Treeview.Heading",
            background=self.colors["primary"],
            foreground=self.colors["text_light"],
            font=("Arial", 16, "bold"),
        )
        style.map(
            "Report.Treeview",
            background=[("selected", "#14b8a6")],
            foreground=[("selected", self.colors["text_light"])],
        )

        columns = ("S.No", "CID", "Name", "Phone", "Address", "Total Liters", "Total Amount")

        self.tree = ttk.Treeview(
        table_container,
        columns=columns,
        show="headings",
        selectmode="browse",
        style="Report.Treeview",
        )

        self.tree.pack(fill="both", expand=True)

        widths = [80, 120, 180, 150, 300, 150, 150]
        for col, width in zip(columns, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")

        vsb = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        # ✅ Keyboard bindings
        self.tree.bind("<Up>", self._on_arrow_up)
        self.tree.bind("<Down>", self._on_arrow_down)


        def on_mouse_wheel(self, event):
            """Handle mouse wheel scrolling"""
            if event.num == 5 or event.delta < 0:
                self.tree.yview_scroll(1, "units")
            elif event.num == 4 or event.delta > 0:
                self.tree.yview_scroll(-1, "units")

    def calculate_total_liters(self):
        """Calculate total liters for each customer from all monthly sheets"""
        for customer in self.all_customers:
            cid = customer.get('CID', '')
            total_liters = 0.0
            total_amount = 0.0
            
            if os.path.exists(self.monthly_sheets_path):
                for filename in os.listdir(self.monthly_sheets_path):
                    if filename.endswith('.xlsx'):
                        filepath = os.path.join(self.monthly_sheets_path, filename)
                        try:
                            wb = load_workbook(filepath, data_only=True)
                            ws = wb.active
                            
                            parts = filename.replace('.xlsx', '').split('_')
                            if len(parts) == 2:
                                month_num = int(parts[1])
                                year = int(parts[0])
                                days_in_month = calendar.monthrange(year, month_num)[1]
                                col_start = 6
                                total_ltr_col = col_start + days_in_month
                                total_amt_col = total_ltr_col + 1
                                
                                for row_idx in range(2, ws.max_row + 1):
                                    cell_cid = ws.cell(row=row_idx, column=2).value
                                    
                                    if str(cell_cid).strip() == str(cid).strip():
                                        ltr_val = ws.cell(row=row_idx, column=total_ltr_col).value
                                        amt_val = ws.cell(row=row_idx, column=total_amt_col).value
                                        
                                        try:
                                            total_liters += float(ltr_val) if ltr_val else 0.0
                                            total_amount += float(amt_val) if amt_val else 0.0
                                        except:
                                            pass
                                        break
                            
                            wb.close()
                        except Exception as e:
                            print(f"Error reading {filename}: {e}")
                            continue
            
            customer['Total_Liters'] = total_liters
            customer['Total_Amount'] = total_amount
                        
    def populate_table(self, data=None):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if data is None:
            data = self.report_data

        try:
            data = sorted(data, key=lambda x: str(x.get("CID", "")).zfill(10))
        except:
            pass

        for idx, customer in enumerate(data, 1):
            tag = "oddrow" if idx % 2 else "evenrow"

            values = (
                str(idx),
                customer.get('CID', ''),
                customer.get('Name', ''),
                customer.get('Phone', ''),
                customer.get('Address', ''),
                f"{float(customer.get('Total_Liters', 0)):.2f}",
                f"₹{float(customer.get('Total_Amount', 0)):.2f}"
            )

            self.tree.insert("", "end", values=values, tags=(tag,))

        # ✅ IMPORTANT: restore focus for arrow keys
        self._focus_first_row()


    def search_customers(self):
        search_term = self.ent_search.get().strip().lower()

        if not search_term:
            self.populate_table(self.report_data)
            self.lbl_status.configure(text=f"Showing {len(self.report_data)} records")
            return

        filtered_data = [
            c for c in self.report_data
            if search_term in str(c.get('CID', '')).lower()
            or search_term in str(c.get('Name', '')).lower()
            or search_term in str(c.get('Phone', '')).lower()
            or search_term in str(c.get('Address', '')).lower()
        ]

        if filtered_data:
            self.populate_table(filtered_data)
            self.lbl_status.configure(text=f"Found {len(filtered_data)} records")
        else:
            messagebox.showinfo(
                "No Results",
                f"No customers found matching '{search_term}'",
                parent=self
            )
            self.lbl_status.configure(text="No results found")


    def clear_search(self):
        self.ent_search.delete(0, "end")
        self.populate_table(self.report_data)
        self.lbl_status.configure(text=f"Showing {len(self.report_data)} records")


    def create_whatsapp_message(self, name, cid, total_ltr, total_amt, month_name, year, days, phone):
        """Create formatted WhatsApp message"""
        # UPDATED: Use config values
        lines = [
            f"🥛 {self.business_name} - Monthly Bill",
            "",
            f"Dear {name},",
            "",
            f"📅 Period: 01 {month_name[:3].upper()} {year} - {days} {month_name[:3].upper()} {year}",
            f"📱 Phone: {phone}",
            "",
            "📊 Delivery Summary:",
            "━━━━━━━━━━━━━━━━━━",
            f"🥛 Total Milk: {total_ltr:.2f} Liters",
            f"💰 Amount Due: ₹{total_amt:.2f}",
            "━━━━━━━━━━━━━━━━━━",
            "",
            f"📞 For queries: {self.contact_number}",
            f"💳 Pay via: {self.payment_info}",
            "",
            "Thank you! 🙏"
        ]
        return "\n".join(lines)

    def download_all_records(self):
        """Download all customer records as Excel"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"All_Customer_Records_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                parent=self
            )
            
            if not file_path:
                return
            
            # Create DataFrame
            data_for_export = []
            for idx, customer in enumerate(self.all_customers, 1):
                data_for_export.append({
                    'S.No': idx,
                    'CID': customer.get('CID', ''),
                    'Name': customer.get('Name', ''),
                    'Phone': customer.get('Phone', ''),
                    'Address': customer.get('Address', ''),
                    'Total Liters': float(customer.get('Total_Liters', 0)),
                    'Total Amount': float(customer.get('Total_Amount', 0)),
                    'Status': customer.get('Status', 'Active')
                })
            
            df = pd.DataFrame(data_for_export)
            
            # Write to Excel with styling
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='All Records')
                
                workbook = writer.book
                worksheet = writer.sheets['All Records']
                
                # Styling
                header_fill = PatternFill(start_color="14b8a6", end_color="14b8a6", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                center_align = Alignment(horizontal='center', vertical='center')
                light_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                
                # Header styling
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                # Apply alternating row colors
                for row in range(2, len(df) + 2):
                    if row % 2 == 0:
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = light_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            messagebox.showinfo(
                "Download Complete",
                f"✅ All records downloaded successfully!\n\nTotal Records: {len(self.all_customers)}\n\nSaved to: {file_path}",
                parent=self
            )
            
        except Exception as e:
            messagebox.showerror("Download Error", f"Failed to download records: {e}", parent=self)

    def download_monthly_report(self):
        try:
            month_name = self.month_names[self.selected_month - 1]
            year = self.selected_year
            days_in_month = calendar.monthrange(year, self.selected_month)[1]

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{month_name}_{year}_Monthly_Report.xlsx",
                parent=self
            )

            if not file_path:
                return

            text_rows = []
            message_rows = []
            link_rows = []

            for idx, customer in enumerate(self.report_data, 1):
                cid = customer.get('CID', '')
                name = customer.get('Name', '')
                phone = str(customer.get('Phone', '')).strip()
                address = customer.get('Address', '')
                total_ltr = float(customer.get('Total_Liters', 0))
                total_amt = float(customer.get('Total_Amount', 0))

                message = self.create_whatsapp_message(
                    name, cid, total_ltr, total_amt,
                    month_name, year, days_in_month, phone
                )

                encoded_message = urllib.parse.quote(message)
                whatsapp_link = f"https://web.whatsapp.com/send?phone={phone}&text={encoded_message}" if phone else ""

                text_rows.append({
                    "S.No": idx,
                    "CID": cid,
                    "Name": name,
                    "Phone": phone,
                    "Address": address,
                    "Total Liters": total_ltr,
                    "Total Amount": total_amt
                })

                message_rows.append({
                    "S.No": idx,
                    "CID": cid,
                    "Name": name,
                    "Phone": phone,
                    "WhatsApp Message": message
                })

                link_rows.append({
                    "S.No": idx,
                    "CID": cid,
                    "Name": name,
                    "Phone": phone,
                    "WhatsApp Link": whatsapp_link
                })

            df_text = pd.DataFrame(text_rows)
            df_message = pd.DataFrame(message_rows)
            df_link = pd.DataFrame(link_rows)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df_text.to_excel(writer, index=False, sheet_name="Text")
                df_message.to_excel(writer, index=False, sheet_name="WhatsApp_Message")
                df_link.to_excel(writer, index=False, sheet_name="WhatsApp_Link")

                # Styling
                center_align = Alignment(horizontal="center", vertical="center")
                left_top_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
                light_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

                # Style all sheets
                for sheet_name, df in [("Text", df_text), ("WhatsApp_Message", df_message), ("WhatsApp_Link", df_link)]:
                    ws = writer.sheets[sheet_name]
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_len = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                max_len = max(max_len, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                    
                    # Apply alternating row colors
                    for row in range(2, len(df) + 2):
                        if row % 2 == 0:
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = light_fill

                # Special styling for WhatsApp_Message sheet
                ws_msg = writer.sheets["WhatsApp_Message"]
                for row in range(2, len(df_message) + 2):
                    ws_msg.row_dimensions[row].height = 220
                    # Center align customer details (columns 1-4)
                    for col in range(1, 5):
                        ws_msg.cell(row=row, column=col).alignment = center_align
                    # Left align and wrap message (column 5)
                    ws_msg.cell(row=row, column=5).alignment = left_top_align

                # Special styling for WhatsApp_Link sheet
                ws_link = writer.sheets["WhatsApp_Link"]
                for row in range(2, len(df_link) + 2):
                    cell = ws_link.cell(row=row, column=5)
                    if cell.value:
                        cell.font = Font(color="0000FF", underline="single")
                        cell.style = "Hyperlink"

            messagebox.showinfo(
                "Download Complete",
                "✅ Monthly report downloaded with 3 sheets:\n\n"
                "1️⃣ Text - Customer data\n"
                "2️⃣ WhatsApp_Message - Formatted messages\n"
                "3️⃣ WhatsApp_Link - Clickable WhatsApp Web links\n\n"
                "📌 Click links in WhatsApp_Link sheet to open Web WhatsApp with pre-filled message.",
                parent=self
            )

        except Exception as e:
            messagebox.showerror("Download Error", str(e), parent=self)

    def show_download_options(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Download Monthly Report")
        dialog.geometry("480x260")
        dialog.transient(self)
        dialog.grab_set()

        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 240
        y = (dialog.winfo_screenheight() // 2) - 130
        dialog.geometry(f"480x260+{x}+{y}")

        main = ctk.CTkFrame(dialog, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=30, pady=30)

        ctk.CTkLabel(
            main,
            text="📥 Download Monthly Report",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=self.colors["primary"]
        ).pack(pady=(0, 25))

        ctk.CTkButton(
            main,
            text="📊 Excel (Text + WhatsApp Message + Link)",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=380,
            height=80,
            corner_radius=16,
            fg_color="#10b981",
            hover_color="#059669",
            command=lambda: (dialog.destroy(), self.download_monthly_report())
        ).pack(pady=10)

        ctk.CTkButton(
            main,
            text="Cancel",
            font=ctk.CTkFont(size=14),
            width=160,
            height=40,
            corner_radius=10,
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=dialog.destroy
        ).pack(pady=(20, 0))


    def debug_check_data(self):
        """Debug function to check if monthly sheets exist and have data"""
        
        
        if os.path.exists(self.monthly_sheets_path):
            files = [f for f in os.listdir(self.monthly_sheets_path) if f.endswith('.xlsx')]
            return
        else:
            print("ERROR: Monthly sheets path does not exist!")
        print("=" * 50)

    def show_year_selection_for_yearly(self):
        """Show year selection screen for yearly report (Dropdown version)"""
        for widget in self.winfo_children():
            widget.destroy()

        # Header
        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        btn_back = ctk.CTkButton(
            header,
            text="Back",
            font=ctk.CTkFont(size=24, weight="bold"),
            width=150,
            height=55,
            corner_radius=15,
            fg_color=self.colors["text_light"],
            text_color=self.colors["primary"],
            hover_color="#e2e8f0",
            command=self.init_menu_ui,
        )
        btn_back.pack(side="left", padx=30, pady=12)

        lbl_title = ctk.CTkLabel(
            header,
            text="📆 Yearly Report",
            font=ctk.CTkFont(size=36, weight="bold"),
            text_color=self.colors["text_light"],
        )
        lbl_title.pack(side="left", padx=20)

        # Content
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=40, pady=40)  # ← CRITICAL FIX

        ctk.CTkLabel(
            content,
            text="Select Year",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=self.colors["text_dark"]
        ).pack(pady=(30, 20))

        # Year dropdown
        year_values = [str(y) for y in range(2020, 2051)]
        self.year_var = ctk.StringVar(value=str(self.current_year))

        year_dropdown = ctk.CTkComboBox(
            content,
            values=year_values,
            variable=self.year_var,
            width=200,
            height=45,
            corner_radius=12,
            font=ctk.CTkFont(size=18, weight="bold")
        )
        year_dropdown.pack(pady=10)

        # View report button
        ctk.CTkButton(
            content,
            text="📊 View Yearly Report",
            font=ctk.CTkFont(size=18, weight="bold"),
            width=300,
            height=60,
            corner_radius=18,
            fg_color="#10b981",
            hover_color="#059669",
            command=lambda: self.show_yearly_data(int(self.year_var.get()))
        ).pack(pady=30)


    def show_yearly_data(self, year):
        """Display yearly data for selected year
        UPDATED: Always fetch latest name/phone/address from main customers file"""
        for widget in self.winfo_children():
            widget.destroy()
        
        self.selected_year = year
        self.view_mode = "yearly"
        
        # Check if any monthly files exist for this year
        yearly_files = []
        if os.path.exists(self.monthly_sheets_path):
            for filename in os.listdir(self.monthly_sheets_path):
                if filename.startswith(f"{year}_") and filename.endswith('.xlsx'):
                    yearly_files.append(filename)
        
        if not yearly_files:
            messagebox.showwarning(
                "No Data Found",
                f"❌ No data found for year {year}",
                parent=self
            )
            self.show_year_selection_for_yearly()
            return
        
        # Load yearly data - aggregate from all monthly files of this year
        try:
            # FIRST: Load latest customer data from main file
            customers_dict = {}
            try:
                df_customers = pd.read_excel(self.customers_file, engine="openpyxl")
                for _, row in df_customers.iterrows():
                    cid = str(row.get('CID', '')).strip()
                    if cid:
                        customers_dict[cid] = {
                            'Name': row.get('Name', ''),
                            'Phone': row.get('Phone', ''),
                            'Address': row.get('Address', '')
                        }
            except Exception as e:
                print(f"Error loading main customers: {e}")
            
            yearly_data = {}
            
            for filename in yearly_files:
                filepath = os.path.join(self.monthly_sheets_path, filename)
                try:
                    wb = load_workbook(filepath, data_only=True)
                    ws = wb.active
                    
                    parts = filename.replace('.xlsx', '').split('_')
                    if len(parts) == 2:
                        month_num = int(parts[1])
                        days_in_month = calendar.monthrange(year, month_num)[1]
                        col_start = 6
                        total_ltr_col = col_start + days_in_month
                        total_amt_col = total_ltr_col + 1
                        
                        for row_idx in range(2, ws.max_row + 1):
                            cid = ws.cell(row=row_idx, column=2).value
                            name_sheet = ws.cell(row=row_idx, column=3).value
                            
                            if not cid or not name_sheet:
                                continue
                            
                            cid_str = str(cid).strip()
                            
                            ltr_val = ws.cell(row=row_idx, column=total_ltr_col).value
                            amt_val = ws.cell(row=row_idx, column=total_amt_col).value
                            
                            try:
                                liters = float(ltr_val) if ltr_val else 0.0
                                amount = float(amt_val) if amt_val else 0.0
                            except:
                                liters = 0.0
                                amount = 0.0
                            
                            if cid_str not in yearly_data:
                                # Get LATEST data from main customers file
                                if cid_str in customers_dict:
                                    name = customers_dict[cid_str]['Name']
                                    phone = customers_dict[cid_str]['Phone']
                                    address = customers_dict[cid_str]['Address']
                                else:
                                    # Fallback for deleted customers
                                    name = name_sheet
                                    phone = ws.cell(row=row_idx, column=4).value or ''
                                    address = ''
                                
                                yearly_data[cid_str] = {
                                    'CID': cid,
                                    'Name': name,
                                    'Phone': phone,
                                    'Address': address,
                                    'Total_Liters': 0.0,
                                    'Total_Amount': 0.0
                                }
                            
                            yearly_data[cid_str]['Total_Liters'] += liters
                            yearly_data[cid_str]['Total_Amount'] += amount
                    
                    wb.close()
                except Exception as e:
                    print(f"Error reading {filename}: {e}")
                    continue
            
            self.report_data = list(yearly_data.values())
            self.report_data = sorted(self.report_data, key=lambda x: str(x.get("CID", "")).zfill(10))
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load yearly data: {e}", parent=self)
            self.show_year_selection_for_yearly()
            return
        
        # Build UI (rest remains the same as before)
        # Header
        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        btn_back = ctk.CTkButton(
            header,
            text="Back",
            font=ctk.CTkFont(size=24, weight="bold"),
            width=150,
            height=55,
            corner_radius=15,
            fg_color=self.colors["text_light"],
            text_color=self.colors["primary"],
            hover_color="#e2e8f0",
            command=self.show_year_selection_for_yearly,
        )
        btn_back.pack(side="left", padx=30, pady=12)
        
        lbl_title = ctk.CTkLabel(
            header,
            text=f"📅 Year {year} Report",
            font=ctk.CTkFont(size=36, weight="bold"),
            text_color=self.colors["text_light"],
        )
        lbl_title.pack(side="left", padx=20)
        
        # Toolbar
        toolbar = ctk.CTkFrame(self, fg_color=self.colors["secondary"], height=120)
        toolbar.pack(fill="x", padx=20, pady=(10, 5))
        toolbar.pack_propagate(False)
        
        # Search frame
        search_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        search_frame.pack(side="left", padx=20, pady=10)
        
        self.ent_search = ctk.CTkEntry(
            search_frame,
            placeholder_text="Search CID, Name, Phone",
            font=ctk.CTkFont(size=16),
            width=350,
            height=45,
            corner_radius=15,
            border_width=2,
            border_color=self.colors["info"],
            fg_color="#ffffff",
            text_color="#000000",
        )
        self.ent_search.pack(side="left", padx=5)
        
        btn_search = ctk.CTkButton(
            search_frame,
            text="Search",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color=self.colors["info"],
            hover_color="#0891b2",
            command=self.search_customers,
        )
        btn_search.pack(side="left", padx=5)
        
        btn_clear = ctk.CTkButton(
            search_frame,
            text="Clear",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=self.clear_search,
        )
        btn_clear.pack(side="left", padx=5)
        
        btn_refresh = ctk.CTkButton(
            search_frame,
            text="Refresh",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=120,
            height=45,
            corner_radius=12,
            fg_color="#f59e0b",
            hover_color="#d97706",
            command=lambda: self.show_yearly_data(self.selected_year)
        )
        btn_refresh.pack(side="left", padx=5)
        
        # Download button
        btn_download = ctk.CTkButton(
            toolbar,
            text="📥 DOWNLOAD",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=200,
            height=55,
            corner_radius=14,
            fg_color="#10b981",
            hover_color="#059669",
            command=self.download_yearly_report
        )
        btn_download.pack(side="right", padx=(10, 20), pady=20)
        
        # Yearly Totals
        totals_frame = ctk.CTkFrame(toolbar, fg_color="#1e293b", corner_radius=15)
        totals_frame.pack(side="right", padx=0, pady=10)
        
        year_total_liters = sum(float(c.get('Total_Liters', 0)) for c in self.report_data)
        year_total_amount = sum(float(c.get('Total_Amount', 0)) for c in self.report_data)
        
        ctk.CTkLabel(
            totals_frame,
            text=f"📊 YEAR {year} TOTALS",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#fbbf24"
        ).pack(padx=20, pady=(10, 5))
        
        ctk.CTkLabel(
            totals_frame,
            text=f"🥛 Total Liters: {year_total_liters:.2f} L",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#10b981"
        ).pack(padx=20, pady=2)
        
        ctk.CTkLabel(
            totals_frame,
            text=f"💰 Total Amount: ₹{year_total_amount:.2f}",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#3b82f6"
        ).pack(padx=20, pady=(2, 10))
        
        # Table
        self.create_report_table()
        self.populate_table(self.report_data)
        
        # Bottom frame
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent", height=70)
        bottom_frame.pack(fill="x", padx=20, pady=10)
        bottom_frame.pack_propagate(False)
        
        self.lbl_status = ctk.CTkLabel(
            bottom_frame,
            text=f"Showing {len(self.report_data)} customers for year {year}",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=self.colors["text_dark"]
        )
        self.lbl_status.pack(side="left", padx=20)

    def download_yearly_report(self):
        """Download yearly report as Excel"""
        try:
            year = self.selected_year
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Year_{year}_Report.xlsx",
                parent=self
            )
            
            if not file_path:
                return
            
            # Create DataFrame
            data_for_export = []
            for idx, customer in enumerate(self.report_data, 1):
                data_for_export.append({
                    'S.No': idx,
                    'CID': customer.get('CID', ''),
                    'Name': customer.get('Name', ''),
                    'Phone': customer.get('Phone', ''),
                    'Address': customer.get('Address', ''),
                    'Total Liters': float(customer.get('Total_Liters', 0)),
                    'Total Amount': float(customer.get('Total_Amount', 0))
                })
            
            df = pd.DataFrame(data_for_export)
            
            # Write to Excel with styling
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=f'Year {year}')
                
                workbook = writer.book
                worksheet = writer.sheets[f'Year {year}']
                
                # Styling
                header_fill = PatternFill(start_color="14b8a6", end_color="14b8a6", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                center_align = Alignment(horizontal='center', vertical='center')
                light_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                
                # Header styling
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                # Apply alternating row colors
                for row in range(2, len(df) + 2):
                    if row % 2 == 0:
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = light_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            year_total_liters = sum(float(c.get('Total_Liters', 0)) for c in self.report_data)
            year_total_amount = sum(float(c.get('Total_Amount', 0)) for c in self.report_data)
            
            messagebox.showinfo(
                "Download Complete",
                f"✅ Year {year} report downloaded successfully!\n\n"
                f"Total Records: {len(self.report_data)}\n"
                f"Total Liters: {year_total_liters:.2f} L\n"
                f"Total Amount: ₹{year_total_amount:.2f}\n\n"
                f"Saved to: {file_path}",
                parent=self
            )
            
        except Exception as e:
            messagebox.showerror("Download Error", f"Failed to download yearly report: {e}", parent=self)