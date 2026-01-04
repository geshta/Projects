import os
import datetime
import pandas as pd
import customtkinter as ctk
from tkinter import ttk, messagebox
import time
import calendar
from openpyxl import load_workbook
import threading
from pathlib import Path
import socket
import random
import pyperclip  # preserves newlines reliably when pasting into WhatsApp
# selenium imports are lazy (inside functions) to avoid failing import if not installed

class MessageTab(ctk.CTkFrame):
    def __init__(self, parent, colors, back_callback, customer_tab):
        super().__init__(parent)
        self.parent = parent
        self.colors = colors
        self.back_callback = back_callback
        self.customer_tab = customer_tab

        # UPDATED: Use dynamic paths from app_config
        from app_config import app_config
        self.app_config = app_config
        
        self.monthly_sheets_path = str(app_config.monthly_sheets_path)
        self.status_folder = str(app_config.status_folder)
        
        # Ensure directories exist
        app_config.monthly_sheets_path.mkdir(parents=True, exist_ok=True)
        app_config.status_folder.mkdir(parents=True, exist_ok=True)

        # State & constants
        self.current_year = datetime.date.today().year
        self.month_names = [
            "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
            "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"
        ]

        self.selected_month = None
        self.selected_year = None
        self.customer_data = []
        self.checkboxes = {}
        self.tree_iid_to_key = {}
        self.failed_list = []
        self.sent_list = []
        self.search_mode = False
        self.previously_selected = set()
        self.viewing_unsent_data = []
        self.viewing_sent_data = []

        # UPDATED: Get business info from config
        self.business_name = app_config.get_business_name()
        self.contact_number = app_config.get_contact_number()
        self.payment_info = app_config.get_payment_info()

        # Selenium placeholders
        self.driver = None
        self.By = None
        self.WebDriverWait = None
        self.EC = None
        self.Keys = None
        self.working_send_strategy = None

        # Sending flags
        self.is_sending = False
        self.send_thread = None
        self.is_paused = False
        self.pause_requested = False

        # UI color randomizer
        self._rand_button_color = self._generate_random_color()
        self._rand_button_text_color = "#ffffff"

        # Build UI
        self.pack_forget()
        self.init_month_selection_ui()

    # ---------------------------
    # Utilities
    # ---------------------------

    def _begin_page_build(self):
        self.pack_forget()
        self.update_idletasks()

    def _end_page_build(self):
        self.pack(fill="both", expand=True)
        self.update_idletasks()



    def _generate_random_color(self):
        r = random.randint(80, 200)
        g = random.randint(70, 200)
        b = random.randint(70, 200)
        return "#{:02x}{:02x}{:02x}".format(r, g, b)
    
    def reload_customers(self):
        """Reload customer data when changes are made in customer tab"""
        try:
            # If we have a month selected and customer data loaded, refresh it
            if hasattr(self, 'selected_month') and self.selected_month:
                if hasattr(self, 'customer_data') and self.customer_data:
                    # Reload the current month's data
                    self.load_customer_list(self.selected_month)
        except Exception as e:
            print(f"Message tab reload error: {e}")


    def check_internet(self, host="8.8.8.8", port=53, timeout=2):
        try:
            socket.setdefaulttimeout(timeout)
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.connect((host, port))
            sock.close()
            return True
        except:
            return False

    def truncate_reason(self, reason, max_words=3):
        try:
            if not reason:
                return ""
            words = str(reason).strip().split()
            return " ".join(words[:max_words])
        except:
            return ""

    def format_phone_number(self, phone):
        phone = str(phone).strip()
        phone = ''.join(filter(str.isdigit, phone))
        if len(phone) == 10:
            return f"91{phone}"
        elif len(phone) == 12 and phone.startswith("91"):
            return phone
        elif len(phone) == 11 and phone.startswith("0"):
            return f"91{phone[1:]}"
        else:
            return None

    # ---------------------------
    # UI Initialization
    # ---------------------------
    def init_month_selection_ui(self):
        for widget in self.winfo_children():
            widget.destroy()

        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        btn_back = ctk.CTkButton(
            header, text="Back", font=ctk.CTkFont(size=24, weight="bold"),
            width=150, height=55, corner_radius=15, fg_color=self.colors["text_light"],
            text_color=self.colors["primary"], hover_color="#e2e8f0",
            command=self.back_callback
        )
        btn_back.pack(side="left", padx=30, pady=12)

        lbl_title = ctk.CTkLabel(
            header, text="📱 WhatsApp Message Management",
            font=ctk.CTkFont(size=36, weight="bold"),
            text_color=self.colors["text_light"]
        )
        lbl_title.pack(side="left", padx=20)

        content_frame = ctk.CTkFrame(self, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=40, pady=20)

        # ----------------------------
        # YEAR SELECTOR (extended to 2050)
        # ----------------------------
        year_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        year_frame.pack(pady=(10, 5))

        ctk.CTkLabel(
            year_frame, text="Select Year:",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=self.colors["text_dark"]
        ).pack(side="left", padx=10)

        # Current year → 2050
        year_values = [str(y) for y in range(2020, 2051)]

        self.year_var = ctk.StringVar(value=str(self.current_year))
        year_selector = ctk.CTkComboBox(
            year_frame,
            values=year_values,
            variable=self.year_var,
            width=150,
            height=40,
            corner_radius=10,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        year_selector.pack(side="left", padx=10)

        # ----------------------------
        # MONTH BUTTONS
        # ----------------------------
        ctk.CTkLabel(
            content_frame,
            text="Select Month to View Customers",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=self.colors["text_dark"]
        ).pack(pady=20)

        btn_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        btn_frame.pack()

        # Detect current system month
        current_month = datetime.date.today().month

        for i, month_name in enumerate(self.month_names, 1):

            # Highlight style for current month
            if i == current_month:
                fg = "#22c55e"          # green highlight
                hover = "#16a34a"
                border_color = "#14532d"
                border_width = 3
                text_color = "#ffffff"
            else:
                fg = self.colors["primary"]
                hover = "#1e40af"
                border_color = "#cbd5f5"
                border_width = 1
                text_color = "#ffffff"

            button = ctk.CTkButton(
                btn_frame,
                text=month_name,
                width=140,
                height=48,
                font=ctk.CTkFont(size=14, weight="bold"),
                corner_radius=14,
                fg_color=fg,
                hover_color=hover,
                border_color=border_color,
                border_width=border_width,
                text_color=text_color,
                command=lambda m=i: self.load_customer_list(m)
            )

            button.grid(
                row=(i - 1) // 4,
                column=(i - 1) % 4,
                padx=10,
                pady=10
            )



    # ---------------------------
    # Load & display customers (sheet -> memory)
    # ---------------------------
    def load_customer_list(self, month):
        """
        Load customers AND manually calculate totals so Excel formula evaluation is not required.
        Also load persisted sent_list for this month so sent_list persists across restarts.
        UPDATED: Reload fresh data from main customers file for latest name/phone/address
        """
        self.selected_month = month
        try:
            self.selected_year = int(self.year_var.get())
        except:
            self.selected_year = datetime.date.today().year

        monthly_path = os.path.join(self.monthly_sheets_path, f"{self.selected_year}_{self.selected_month:02d}.xlsx")

        if not os.path.exists(monthly_path):
            messagebox.showerror("Missing File",
                                f"No monthly sheet found for {self.month_names[month-1]} {self.selected_year}.",
                                parent=self)
            return

        try:
            # FIRST: Load current customer data from main customers file
            customers_dict = {}
            try:
                df_customers = pd.read_excel(self.customer_tab.excel_file, engine="openpyxl")
                for _, row in df_customers.iterrows():
                    cid = str(row.get('CID', '')).strip()
                    if cid:
                        customers_dict[cid] = {
                            'Name': row.get('Name', ''),
                            'Phone': row.get('Phone', ''),
                            'Address': row.get('Address', '')
                        }
            except Exception as e:
                print(f"Error loading main customers: {e}")
            
            # THEN: Load monthly sheet data
            wb = load_workbook(monthly_path, data_only=False)
            ws = wb.active

            days_in_month = calendar.monthrange(self.selected_year, self.selected_month)[1]
            col_start = 6
            col_end = col_start + days_in_month - 1

            self.customer_data = []

            for row in range(2, ws.max_row + 1):
                cid = ws.cell(row=row, column=2).value
                name_sheet = ws.cell(row=row, column=3).value
                phone_sheet = ws.cell(row=row, column=4).value

                if not cid or not name_sheet:
                    continue
                
                cid_str = str(cid).strip()

                # Get LATEST data from main customers file
                if cid_str in customers_dict:
                    name = customers_dict[cid_str]['Name']
                    phone = customers_dict[cid_str]['Phone']
                    address = customers_dict[cid_str]['Address']
                else:
                    # Fallback to sheet data if customer deleted
                    name = name_sheet
                    phone = phone_sheet
                    address = ""

                # Manual total computation
                total_ltr = 0
                for col in range(col_start, col_end + 1):
                    val = ws.cell(row=row, column=col).value
                    try:
                        total_ltr += float(val) if val else 0
                    except:
                        pass

                # Fetch rate
                try:
                    rate = float(ws["E1"].value)
                except:
                    rate = 0

                total_amt = total_ltr * rate

                self.customer_data.append({
                    "sno": row-1,
                    "CID": cid,
                    "Name": name,
                    "Phone": phone,
                    "address": address,
                    "Total_Ltr": total_ltr,
                    "Total_Amt": total_amt
                })

            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data:\n{e}")
            return

        # Load persisted sent_list
        self.load_sent_list()

        # Show customer table
        self.show_customer_table()

    # message_tab.py  — Part 2/3 (Updated)

    # ---------------------------
    # Table UI
    # ---------------------------
    def show_customer_table(self):
        """Build the table UI with arrow key navigation support."""
        for widget in self.winfo_children():
            widget.destroy()

        header = ctk.CTkFrame(self, fg_color=self.colors["primary"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        btn_back = ctk.CTkButton(
            header, text="Back", font=ctk.CTkFont(size=24, weight="bold"),
            width=150, height=55, corner_radius=15, fg_color=self.colors["text_light"],
            text_color=self.colors["primary"], hover_color="#e2e8f0",
            command=self.init_month_selection_ui
        )
        btn_back.pack(side="left", padx=30, pady=12)

        lbl_title = ctk.CTkLabel(
            header, text=f"📱 {self.month_names[self.selected_month-1]} {self.selected_year} - Messages",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=self.colors["text_light"]
        )
        lbl_title.pack(side="left", padx=20)

        toolbar = ctk.CTkFrame(self, fg_color=self.colors["secondary"], height=70)
        toolbar.pack(fill="x", padx=20, pady=(10, 5))
        toolbar.pack_propagate(False)

        # refresh
        self.btn_refresh = ctk.CTkButton(
            toolbar, text="🔄 Refresh", font=ctk.CTkFont(size=16, weight="bold"),
            width=120, height=45, corner_radius=12, fg_color=self._rand_button_color,
            text_color=self._rand_button_text_color, hover_color=self._rand_button_color,
            command=self.refresh_data
        )
        self.btn_refresh.pack(side="left", padx=10, pady=10)

        # select all
        # select all + range selector
        select_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        select_frame.pack(side="left", padx=10, pady=15)

        self.select_all_var = ctk.BooleanVar(value=False)
        chk_select_all = ctk.CTkCheckBox(
            select_frame, text="Select All", variable=self.select_all_var,
            font=ctk.CTkFont(size=18, weight="bold"), command=self.toggle_select_all,
            checkbox_width=25, checkbox_height=25
        )
        chk_select_all.pack(side="left", padx=5)

        # Range selector button
        btn_range = ctk.CTkButton(
            select_frame,
            text="🎯 Range",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=100,
            height=45,
            corner_radius=12,
            fg_color="#8b5cf6",
            hover_color="#7c3aed",
            command=self.show_range_selector
        )
        btn_range.pack(side="left", padx=5)

        # search
        self.ent_search = ctk.CTkEntry(
            toolbar, placeholder_text="Search CID, Name, Phone, Address",
            font=ctk.CTkFont(size=16), width=350, height=45, corner_radius=15,
            border_width=2, border_color=self.colors["info"], fg_color="#ffffff", text_color="#000000"
        )
        self.ent_search.pack(side="left", padx=10)
        btn_search = ctk.CTkButton(
            toolbar, text="Search", font=ctk.CTkFont(size=18, weight="bold"),
            width=80, height=45, corner_radius=15, fg_color=self.colors["info"],
            hover_color="#0891b2", command=self.search_customers
        )
        btn_search.pack(side="left", padx=10)

        btn_clear_search = ctk.CTkButton(
            toolbar, text="Clear Search", font=ctk.CTkFont(size=18, weight="bold"),
            width=100, height=45, corner_radius=15, fg_color="#6b7280",
            hover_color="#4b5563", command=self.clear_search
        )
        btn_clear_search.pack(side="left", padx=5)

        self.lbl_selected = ctk.CTkLabel(
            toolbar, text=f"Selected: 0/{len(self.customer_data)} customers",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=self.colors["text_dark"]
        )
        self.lbl_selected.pack(side="right", padx=20)

        # customer list area
        frame_list = ctk.CTkFrame(self, fg_color=self.colors["text_light"])
        frame_list.pack(fill="both", expand=True, padx=25, pady=(0, 10))

        table_container = ctk.CTkFrame(frame_list, fg_color="transparent")
        table_container.pack(fill="both", expand=True)

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure(
            "Custom.Treeview", background="#ffffff", foreground=self.colors["text_dark"],
            fieldbackground="#ffffff", borderwidth=0, rowheight=46, font=("Arial", 14)
        )
        style.configure(
            "Custom.Treeview.Heading", background=self.colors["primary"], foreground=self.colors["text_light"],
            borderwidth=0, relief="flat", font=("Arial", 16, "bold")
        )
        style.map("Custom.Treeview", 
                background=[("selected", "#14b8a6")], 
                foreground=[("selected", self.colors["text_light"])])

        self.columns = ("chk", "sno", "cid", "name", "phone", "address", "ltr", "amt")
        self.tree = ttk.Treeview(table_container, columns=self.columns, show="headings", 
                                selectmode="browse", style="Custom.Treeview")
        self.tree.pack(fill="both", expand=True)

        widths = [80, 80, 100, 200, 150, 350, 120, 120]
        headings = ["Select", "S.No", "CID", "Name", "Phone", "Address", "Total Ltr", "Total Amt"]
        for col, w, head in zip(self.columns, widths, headings):
            self.tree.heading(col, text=head)
            self.tree.column(col, width=w, anchor="center")

        vsb = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree.tag_configure("oddrow", background="#e0f2fe")
        self.tree.tag_configure("evenrow", background="#ffffff")
        self.tree.tag_configure("highlight", background="#fbbf24", foreground="#000000")

        # init internal maps
        self.checkboxes = {}
        self.tree_iid_to_key = {}
        self.populate_table()

        # bindings - UPDATED WITH ARROW KEYS
        self.tree.bind("<Button-1>", self.on_tree_click)
        self.tree.bind("<MouseWheel>", self.on_mouse_wheel)
        self.tree.bind("<Button-4>", self.on_mouse_wheel)
        self.tree.bind("<Button-5>", self.on_mouse_wheel)
        
        # Arrow key navigation
        self.tree.bind("<Up>", self.on_arrow_up)
        self.tree.bind("<Down>", self.on_arrow_down)
        self.tree.bind("<space>", self.on_space_toggle)  # Space to toggle checkbox
        self.tree.bind("<Return>", self.on_space_toggle)  # Enter to toggle checkbox

        bottom_frame = ctk.CTkFrame(self, fg_color="transparent", height=80)
        bottom_frame.pack(fill="x", padx=20, pady=(5, 15))
        bottom_frame.pack_propagate(False)

        self.lbl_estimate = ctk.CTkLabel(bottom_frame, text="Estimated Time: ~0 minutes",
                                        font=ctk.CTkFont(size=16), text_color=self.colors["text_dark"])
        self.lbl_estimate.pack(side="left", padx=20)

        btn_status = ctk.CTkButton(bottom_frame, text="📊 See Status", font=ctk.CTkFont(size=16, weight="bold"),
                                width=150, height=50, corner_radius=12, fg_color="#8b5cf6",
                                text_color=self.colors["text_light"], hover_color="#7c3aed",
                                command=self.show_status_options)
        btn_status.pack(side="right", padx=10)

        self.btn_send = ctk.CTkButton(bottom_frame, text="📱 Send to WhatsApp", font=ctk.CTkFont(size=18, weight="bold"),
                                    width=200, height=50, corner_radius=12, fg_color=self.colors["success"],
                                    text_color=self.colors["text_light"], hover_color="#059669",
                                    command=self.send_messages)
        self.btn_send.pack(side="right", padx=10)

        # Set focus to tree for arrow key navigation
        self.tree.focus_set()

    # ---------------------------
    # Table population & mapping
    # ---------------------------
    def populate_table(self, filtered_data=None):
        """Populate tree and maintain stable mapping between iid and customer key."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.checkboxes.clear()
        self.tree_iid_to_key.clear()

        data_to_show = filtered_data if filtered_data is not None else (self.customer_data or [])

        for idx, customer in enumerate(data_to_show):
            # stable key: prefer CID else fallback unique
            cid_val = customer.get('CID') or customer.get('cid') or ""
            cid_key = str(cid_val).strip() if cid_val else f"idx_{idx}_{int(time.time()*1000) % 100000}"

            var = ctk.BooleanVar(value=(cid_key in self.previously_selected))
            self.checkboxes[cid_key] = {'var': var, 'data': customer}

            tag = "oddrow" if idx % 2 else "evenrow"

            # safe numeric parsing
            total_ltr = customer.get('Total_Ltr', customer.get('total_ltr', 0) or 0)
            total_amt = customer.get('Total_Amt', customer.get('total_amt', 0) or 0)
            try:
                total_ltr_f = float(total_ltr)
            except:
                total_ltr_f = 0.0
            try:
                total_amt_f = float(total_amt)
            except:
                total_amt_f = 0.0

            values = (
                "☑" if var.get() else "☐",
                customer.get('sno', idx+1),
                cid_val,
                customer.get('Name', customer.get('name', '')),
                customer.get('Phone', customer.get('phone', '')),
                customer.get('address', ''),
                f"{total_ltr_f:.2f}",
                f"₹{total_amt_f:.2f}"
            )
            iid = self.tree.insert("", "end", values=values, tags=(tag,))
            self.tree_iid_to_key[iid] = cid_key

        self.update_selected_count()
        self.update_checkbox_display()
    
    def on_arrow_up(self, event):
        """Handle Up arrow key - move to previous row"""
        try:
            current_item = self.tree.focus()
            if not current_item:
                # Select first item if nothing selected
                items = self.tree.get_children()
                if items:
                    self.tree.focus(items[0])
                    self.tree.selection_set(items[0])
                    self.highlight_current_row()
                return "break"
            
            # Get previous item
            prev_item = self.tree.prev(current_item)
            if prev_item:
                self.tree.focus(prev_item)
                self.tree.selection_set(prev_item)
                self.tree.see(prev_item)
                self.highlight_current_row()
            
            return "break"  # Prevent default behavior
        except:
            return "break"

    def on_arrow_down(self, event):
        """Handle Down arrow key - move to next row"""
        try:
            current_item = self.tree.focus()
            if not current_item:
                # Select first item if nothing selected
                items = self.tree.get_children()
                if items:
                    self.tree.focus(items[0])
                    self.tree.selection_set(items[0])
                    self.highlight_current_row()
                return "break"
            
            # Get next item
            next_item = self.tree.next(current_item)
            if next_item:
                self.tree.focus(next_item)
                self.tree.selection_set(next_item)
                self.tree.see(next_item)
                self.highlight_current_row()
            
            return "break"  # Prevent default behavior
        except:
            return "break"

    def on_space_toggle(self, event):
        """Handle Space/Enter key - toggle checkbox for current row"""
        try:
            current_item = self.tree.focus()
            if not current_item:
                return "break"
            
            # Get CID from tree_iid_to_key mapping
            cid_key = self.tree_iid_to_key.get(current_item)
            if not cid_key:
                # Fallback to values
                vals = self.tree.item(current_item).get('values', [])
                cid_key = str(vals[2]) if len(vals) >= 3 else None
            
            if cid_key and cid_key in self.checkboxes:
                # Toggle checkbox
                current = self.checkboxes[cid_key]['var'].get()
                self.checkboxes[cid_key]['var'].set(not current)
                if not current:
                    self.previously_selected.add(cid_key)
                else:
                    self.previously_selected.discard(cid_key)
                self.update_checkbox_display()
                self.update_selected_count()
            
            return "break"
        except:
            return "break"

    def highlight_current_row(self):
        """Apply highlight to currently focused row"""
        try:
            # Remove highlight from all rows
            for item in self.tree.get_children():
                current_tags = list(self.tree.item(item, 'tags'))
                # Remove highlight tag if present
                if 'highlight' in current_tags:
                    current_tags.remove('highlight')
                self.tree.item(item, tags=current_tags)
            
            # Add highlight to focused row
            current_item = self.tree.focus()
            if current_item:
                current_tags = list(self.tree.item(current_item, 'tags'))
                if 'highlight' not in current_tags:
                    current_tags.append('highlight')
                self.tree.item(current_item, tags=tuple(current_tags))
        except:
            pass

    def update_checkbox_display(self):
        """Refresh first-column checkbox icons based on internal boolean vars."""
        for iid, cid_key in self.tree_iid_to_key.items():
            cb = self.checkboxes.get(cid_key)
            if not cb:
                continue
            current_icon = "☑" if cb['var'].get() else "☐"
            vals = list(self.tree.item(iid, 'values'))
            if not vals:
                continue
            vals[0] = current_icon
            self.tree.item(iid, values=vals)

    def update_selected_count(self):
        selected = sum(1 for cb in self.checkboxes.values() if cb['var'].get())
        total = len(self.checkboxes)
        try:
            self.lbl_selected.configure(text=f"Selected: {selected}/{total} customers")
        except:
            pass
        est_minutes = max(1, (selected * 9) // 60)
        try:
            self.lbl_estimate.configure(text=f"Estimated Time: ~{est_minutes} minutes")
        except:
            pass

    def on_tree_click(self, event):
        """Identify cell click and toggle checkbox when first column clicked."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if item and col == "#1":
            cid_key = self.tree_iid_to_key.get(item)
            # fallback to CID cell if mapping missing
            if not cid_key:
                vals = self.tree.item(item).get('values', [])
                cid_key = str(vals[2]) if len(vals) >= 3 else None
            if cid_key and cid_key in self.checkboxes:
                current = self.checkboxes[cid_key]['var'].get()
                self.checkboxes[cid_key]['var'].set(not current)
                if not current:
                    self.previously_selected.add(cid_key)
                else:
                    self.previously_selected.discard(cid_key)
                self.update_checkbox_display()
                self.update_selected_count()

    def on_mouse_wheel(self, event):
        """Handle touchpad and mouse wheel scroll events safely."""
        try:
            if not hasattr(self, 'tree') or not self.tree or not getattr(self.tree, 'winfo_exists', lambda: False)():
                return
            if hasattr(event, 'delta'):
                delta = event.delta
            elif hasattr(event, 'num'):
                # Button-4/5 emulation
                delta = 120 if event.num == 4 else -120
            else:
                delta = 0
            if delta < 0:
                self.tree.yview_scroll(1, "units")
            else:
                self.tree.yview_scroll(-1, "units")
        except Exception:
            return

    def toggle_select_all(self):
        select_state = self.select_all_var.get()
        for cid_key, obj in self.checkboxes.items():
            try:
                obj['var'].set(select_state)
                if select_state:
                    self.previously_selected.add(cid_key)
                else:
                    self.previously_selected.discard(cid_key)
            except:
                continue
        self.update_checkbox_display()
        self.update_selected_count()

    # ---------------------------
    # Search & clear
    # ---------------------------
    def search_customers(self):
        search_term = self.ent_search.get().strip().lower()

        if not search_term:
            self.search_mode = False
            self.populate_table()
            self._focus_first_row()
            return

        filtered = [
            c for c in self.customer_data
            if search_term in str(c.get('CID', c.get('cid', ''))).lower()
            or search_term in str(c.get('Name', c.get('name', ''))).lower()
            or search_term in str(c.get('Phone', c.get('phone', ''))).lower()
            or search_term in str(c.get('address', '')).lower()
        ]

        if filtered:
            self.search_mode = True
            self.populate_table(filtered)
            self._focus_first_row()
        else:
            messagebox.showinfo(
                "No Results",
                f"No customers found matching '{search_term}'",
                parent=self
            )
            self.search_mode = False


    def clear_search(self):
        try:
            self.ent_search.delete(0, "end")
        except:
            pass

        self.search_mode = False
        self.populate_table()
        self._focus_first_row()
        self.update_selected_count()

    def _focus_first_row(self):
        try:
            self.tree.focus_set()
            items = self.tree.get_children()
            if items:
                self.tree.focus(items[0])
                self.tree.selection_set(items[0])
                self.tree.see(items[0])
                self.highlight_current_row()
        except:
            pass


    # ---------------------------
    # Status views
    # ---------------------------
    def show_status_options(self):
        status_window = ctk.CTkToplevel(self)
        status_window.title("View Status")
        status_window.geometry("400x250")
        status_window.transient(self)
        status_window.grab_set()

        ctk.CTkLabel(status_window, text="What would you like to view?", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=30)

        def show_sent():
            status_window.destroy()
            self.show_sent_messages()

        def show_unsent():
            status_window.destroy()
            self.show_unsent_messages()

        btn_frame = ctk.CTkFrame(status_window, fg_color="transparent")
        btn_frame.pack(pady=20)

        ctk.CTkButton(btn_frame, text="✅ View Sent Messages", font=ctk.CTkFont(size=16, weight="bold"),
                      width=250, height=50, corner_radius=12, fg_color=self.colors["success"],
                      hover_color="#059669", command=show_sent).pack(pady=10)

        ctk.CTkButton(btn_frame, text="❌ View Unsent Messages", font=ctk.CTkFont(size=16, weight="bold"),
                      width=250, height=50, corner_radius=12, fg_color=self.colors["warning"],
                      hover_color="#d97706", command=show_unsent).pack(pady=10)

    def show_sent_messages(self):
        for widget in self.winfo_children():
            widget.destroy()

        header = ctk.CTkFrame(self, fg_color=self.colors["success"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        btn_back = ctk.CTkButton(header, text="Back", font=ctk.CTkFont(size=24, weight="bold"), width=150, height=55,
                                corner_radius=15, fg_color=self.colors["text_light"], text_color=self.colors["success"],
                                hover_color="#e2e8f0", command=self.show_customer_table)
        btn_back.pack(side="left", padx=30, pady=12)

        ctk.CTkLabel(header, text="✅ Successfully Sent Messages", font=ctk.CTkFont(size=32, weight="bold"),
                     text_color=self.colors["text_light"]).pack(side="left", padx=20)

        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=40, pady=20)

        ctk.CTkLabel(content, text=f"Total Sent: {len(self.sent_list)}", font=ctk.CTkFont(size=24, weight="bold"),
                     text_color=self.colors["success"]).pack(pady=20)

        if not self.sent_list:
            ctk.CTkLabel(content, text="No sent messages yet.", font=ctk.CTkFont(size=18),
                         text_color="#6b7280").pack(pady=50)
            return

        table_container = ctk.CTkFrame(content, fg_color="transparent")
        table_container.pack(fill="both", expand=True)

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure("Sent.Treeview", background="#ffffff", foreground=self.colors["text_dark"], fieldbackground="#ffffff",
                        borderwidth=0, rowheight=48, font=("Arial", 14))
        style.configure("Sent.Treeview.Heading", background=self.colors["success"], foreground=self.colors["text_light"],
                        borderwidth=0, font=("Arial", 16, "bold"))

        columns = ("CID", "Name", "Phone", "Total Ltr", "Total Amt")
        tree = ttk.Treeview(table_container, columns=columns, show="headings", selectmode="none", style="Sent.Treeview")
        tree.pack(fill="both", expand=True)

        widths = [120, 200, 150, 120, 120]
        for col, w in zip(columns, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center")

        vsb = ttk.Scrollbar(table_container, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        tree.tag_configure("oddrow", background="#e0f2fe")
        tree.tag_configure("evenrow", background="#ffffff")

        for idx, cust in enumerate(self.sent_list):
            tag = "oddrow" if idx % 2 else "evenrow"
            vals = (
                cust.get('CID', cust.get('cid', '')),
                cust.get('Name', cust.get('name', '')),
                cust.get('Phone', cust.get('phone', '')),
                f"{float(cust.get('Total_Ltr', 0)):.2f}",
                f"₹{float(cust.get('Total_Amt', 0)):.2f}"
            )
            tree.insert("", "end", values=vals, tags=(tag,))

    # ---------------------------
    # Unsent view + retry
    # ---------------------------
    def show_unsent_messages(self):
        unsent_file = self.get_unsent_file_path()
        if not unsent_file or not os.path.exists(unsent_file):
            messagebox.showinfo("No Unsent Messages", "No unsent messages found for this month.", parent=self)
            self.show_customer_table()
            return

        try:
            df_unsent = pd.read_excel(unsent_file, engine="openpyxl")
            if 'Reason' in df_unsent.columns:
                df_unsent['Reason'] = df_unsent['Reason'].fillna('').apply(lambda r: self.truncate_reason(r, 3))
            self.viewing_unsent_data = df_unsent.to_dict('records')
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load unsent list: {e}", parent=self)
            self.show_customer_table()
            return

        for widget in self.winfo_children():
            widget.destroy()

        header = ctk.CTkFrame(self, fg_color=self.colors["warning"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        btn_back = ctk.CTkButton(header, text="Back", font=ctk.CTkFont(size=24, weight="bold"),
                                width=150, height=55, corner_radius=15, fg_color=self.colors["text_light"],
                                text_color=self.colors["warning"], hover_color="#e2e8f0", command=self.show_customer_table)
        btn_back.pack(side="left", padx=30, pady=12)

        ctk.CTkLabel(header, text="❌ Failed Messages - Retry", font=ctk.CTkFont(size=32, weight="bold"),
                     text_color=self.colors["text_light"]).pack(side="left", padx=20)

        # toolbar
        toolbar = ctk.CTkFrame(self, fg_color=self.colors["secondary"], height=70)
        toolbar.pack(fill="x", padx=20, pady=(10, 5))
        toolbar.pack_propagate(False)

        self.unsent_select_all_var = ctk.BooleanVar(value=False)
        chk_all = ctk.CTkCheckBox(toolbar, text="Select All", variable=self.unsent_select_all_var,
                                  font=ctk.CTkFont(size=18, weight="bold"), command=self.toggle_unsent_select_all,
                                  checkbox_width=25, checkbox_height=25)
        chk_all.pack(side="left", padx=20, pady=15)

        self.unsent_lbl_selected = ctk.CTkLabel(toolbar, text=f"Selected: 0/{len(self.viewing_unsent_data)} customers",
                                                font=ctk.CTkFont(size=16, weight="bold"), text_color=self.colors["text_dark"])
        self.unsent_lbl_selected.pack(side="right", padx=20)

        frame_list = ctk.CTkFrame(self, fg_color=self.colors["text_light"])
        frame_list.pack(fill="both", expand=True, padx=25, pady=(0, 10))

        table_container = ctk.CTkFrame(frame_list, fg_color="transparent")
        table_container.pack(fill="both", expand=True)

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure("Unsent.Treeview", background="#ffffff", foreground=self.colors["text_dark"],
                        fieldbackground="#ffffff", borderwidth=0, rowheight=48, font=("Arial", 14))
        style.configure("Unsent.Treeview.Heading", background=self.colors["warning"], foreground=self.colors["text_light"],
                        borderwidth=0, font=("Arial", 16, "bold"))

        self.unsent_tree = ttk.Treeview(table_container, columns=("sel", "cid", "name", "phone", "ltr", "amt", "reason"),
                                        show="headings", selectmode="none", style="Unsent.Treeview")
        self.unsent_tree.pack(fill="both", expand=True)

        columns_unsent = [("sel", 80, "Select"), ("cid", 120, "CID"), ("name", 200, "Name"), ("phone", 130, "Phone"),
                          ("ltr", 120, "Total Ltr"), ("amt", 120, "Total Amt"), ("reason", 250, "Reason")]
        for col, w, head in columns_unsent:
            self.unsent_tree.heading(col, text=head)
            self.unsent_tree.column(col, width=w, anchor="center")

        vsb = ttk.Scrollbar(table_container, orient="vertical", command=self.unsent_tree.yview)
        hsb = ttk.Scrollbar(table_container, orient="horizontal", command=self.unsent_tree.xview)
        self.unsent_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.unsent_tree.tag_configure("oddrow", background="#e0f2fe")
        self.unsent_tree.tag_configure("evenrow", background="#ffffff")

        # populate unsent_map to track selection
        self.unsent_checkboxes = {}
        for idx, rec in enumerate(self.viewing_unsent_data):
            cid_key = rec.get('CID', rec.get('cid', f"u_{idx}"))
            var = ctk.BooleanVar(value=False)
            self.unsent_checkboxes[cid_key] = {'var': var, 'data': rec}
            tag = "oddrow" if idx % 2 else "evenrow"
            reason = rec.get('Reason', '').strip()

            if not reason:
                reason_display = "Not selected"
            elif "invalid" in reason.lower():
                reason_display = "Invalid number"
            elif "failed" in reason.lower():
                reason_display = "Failed - retry"
            else:
                reason_display = reason

            vals = ("☐", rec.get('CID', ''), rec.get('Name', ''), rec.get('Phone', ''), f"{float(rec.get('Total_Ltr', 0)):.2f}",
                    f"₹{float(rec.get('Total_Amt', 0)):.2f}", reason_display)
            self.unsent_tree.insert("", "end", values=vals, tags=(tag,))

        self.unsent_tree.bind("<Button-1>", self.on_unsent_tree_click)
        self.unsent_tree.bind("<MouseWheel>", self.on_mouse_wheel)
        self.unsent_tree.bind("<Button-4>", self.on_mouse_wheel)
        self.unsent_tree.bind("<Button-5>", self.on_mouse_wheel)

        bottom_frame = ctk.CTkFrame(self, fg_color="transparent", height=80)
        bottom_frame.pack(fill="x", padx=20, pady=(5, 15))
        bottom_frame.pack_propagate(False)

        self.unsent_lbl_estimate = ctk.CTkLabel(bottom_frame, text="Estimated Time: ~0 minutes", font=ctk.CTkFont(size=16),
                                               text_color=self.colors["text_dark"])
        self.unsent_lbl_estimate.pack(side="left", padx=20)

        self.unsent_btn_send = ctk.CTkButton(bottom_frame, text="📱 Retry Failed Messages", font=ctk.CTkFont(size=18, weight="bold"),
                                             width=250, height=50, corner_radius=12, fg_color=self.colors["warning"],
                                             text_color=self.colors["text_light"], hover_color="#d97706", command=self.retry_unsent_messages)
        self.unsent_btn_send.pack(side="right", padx=10)

    def on_unsent_tree_click(self, event):
        region = self.unsent_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        item = self.unsent_tree.identify_row(event.y)
        col = self.unsent_tree.identify_column(event.x)
        if item and col == "#1":
            vals = self.unsent_tree.item(item).get('values', [])
            cid = vals[1] if len(vals) > 1 else None
            if cid and cid in self.unsent_checkboxes:
                cur = self.unsent_checkboxes[cid]['var'].get()
                self.unsent_checkboxes[cid]['var'].set(not cur)
                self.update_unsent_checkbox_display()
                self.update_unsent_selected_count()

    def update_unsent_checkbox_display(self):
        for item in self.unsent_tree.get_children():
            vals = list(self.unsent_tree.item(item).get('values', []))
            if len(vals) > 1:
                cid = vals[1]
                if cid in self.unsent_checkboxes:
                    vals[0] = "☑" if self.unsent_checkboxes[cid]['var'].get() else "☐"
                    self.unsent_tree.item(item, values=vals)

    def toggle_unsent_select_all(self):
        state = self.unsent_select_all_var.get()
        for key in self.unsent_checkboxes:
            self.unsent_checkboxes[key]['var'].set(state)
        self.update_unsent_checkbox_display()
        self.update_unsent_selected_count()

    def update_unsent_selected_count(self):
        selected = sum(1 for cb in self.unsent_checkboxes.values() if cb['var'].get())
        total = len(self.unsent_checkboxes)
        try:
            self.unsent_lbl_selected.configure(text=f"Selected: {selected}/{total} customers")
            est_minutes = max(1, (selected * 9) // 60)
            self.unsent_lbl_estimate.configure(text=f"Estimated Time: ~{est_minutes} minutes")
        except:
            pass

    def retry_unsent_messages(self):
        if not self.check_internet():
            messagebox.showerror("No Internet", "Check your Internet", parent=self)
            return
        
        selected_customers = [cb['data'] for cb in self.unsent_checkboxes.values() if cb['var'].get()]
        if not selected_customers:
            messagebox.showwarning("No Selection", "Please select at least one customer.", parent=self)
            return
        
        confirm = messagebox.askyesno("Confirm Retry", 
                                    f"📱 Retry WhatsApp messages to {len(selected_customers)} customers?\n\n"
                                    "⚠️ Do NOT close the Chrome window during sending", parent=self)
        if not confirm:
            return
        
        # Change button to "Initializing..."
        try:
            self.unsent_btn_send.configure(state="disabled", text="⏳ Initializing...")
        except:
            pass
        
        self.is_sending = True
        self.is_paused = False
        self.pause_requested = False
        self.failed_list = []
        
        # Load existing sent_list
        if not hasattr(self, 'sent_list') or self.sent_list is None:
            self.sent_list = []
        self.load_sent_list()
        
        self.send_thread = threading.Thread(target=self.send_messages_thread, args=(selected_customers, True), daemon=False)
        self.send_thread.start()

    # ---------------------------
    # Selenium init & send
    # ---------------------------
    def get_sent_file_path(self):
        """Return the path for this month's sent file in Status/{Month}_{Year}/ folder."""
        if not self.selected_month or not self.selected_year:
            return None
        month_name = self.month_names[self.selected_month-1].title()
        month_folder = os.path.join(self.status_folder, f"{month_name}_{self.selected_year}")
        os.makedirs(month_folder, exist_ok=True)
        return os.path.join(month_folder, f"Sent_{month_name}_{self.selected_year}.xlsx")

    def load_sent_list(self):
        """
        Load persistent sent_list for the current selected month/year.
        If file missing, leaves self.sent_list as-is (usually empty or already-in-memory).
        """
        sent_file = self.get_sent_file_path()
        if not sent_file or not os.path.exists(sent_file):
            if not hasattr(self, 'sent_list'):
                self.sent_list = []
            return
        
        try:
            df = pd.read_excel(sent_file, engine="openpyxl")
            # Ensure we have list of dicts with required keys
            loaded = []
            for _, r in df.iterrows():
                # Get total amount properly
                total_amt = r.get("Total_Amt") or r.get("total_amt") or r.get("Total_RT") or 0
                try:
                    total_amt = float(total_amt) if total_amt else 0.0
                except:
                    total_amt = 0.0
                
                loaded.append({
                    "CID": str(r.get("CID", "")).strip(),
                    "Name": r.get("Name", ""),
                    "Phone": r.get("Phone", ""),
                    "Total_Ltr": float(r.get("Total_Ltr", 0) or 0),
                    "Total_Amt": total_amt,
                })
            
            # Initialize sent_list if not exists
            if not hasattr(self, 'sent_list'):
                self.sent_list = []
            
            # Avoid duplicates: merge with existing
            existing_cids = set([str(c.get("CID", "")).strip() for c in self.sent_list])
            for item in loaded:
                if item.get("CID") not in existing_cids:
                    self.sent_list.append(item)
                    existing_cids.add(item.get("CID"))
        except Exception as e:
            print(f"Failed to load sent list: {e}")
            if not hasattr(self, 'sent_list'):
                self.sent_list = []

    def save_sent_list(self):
        """
        Save current self.sent_list to per-month sent file in month-specific folder.
        Overwrites file with current state.
        """
        try:
            sent_file = self.get_sent_file_path()
            if not sent_file:
                return
            
            if self.sent_list:
                # Ensure all records have proper Total_Amt
                for customer in self.sent_list:
                    if 'Total_Amt' not in customer or customer['Total_Amt'] == 0:
                        total_amt = customer.get('Total_RT') or customer.get('total_amt') or 0
                        try:
                            customer['Total_Amt'] = float(total_amt) if total_amt else 0.0
                        except:
                            customer['Total_Amt'] = 0.0
                
                df = pd.DataFrame(self.sent_list)
                # Ensure column order
                column_order = ["CID", "Name", "Phone", "Total_Ltr", "Total_Amt"]
                df = df[column_order]
                df.to_excel(sent_file, index=False, engine="openpyxl", 
                        sheet_name=self.month_names[self.selected_month-1] if self.selected_month else "Sent")
            else:
                # Create empty file with headers
                df = pd.DataFrame(columns=["CID", "Name", "Phone", "Total_Ltr", "Total_Amt"])
                df.to_excel(sent_file, index=False, engine="openpyxl")
        except Exception as e:
            print(f"Error saving sent list: {e}")

    def init_selenium(self):
        """Initialize Selenium with auto-updating Chrome driver"""
        try:
            import undetected_chromedriver as uc
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.common.keys import Keys

            self.By = By
            self.WebDriverWait = WebDriverWait
            self.EC = EC
            self.Keys = Keys

            # Check if driver already exists and is working
            if hasattr(self, "driver") and self.driver:
                try:
                    _ = self.driver.current_url
                    return True
                except:
                    try: 
                        self.driver.quit()
                    except: 
                        pass
                    self.driver = None

            # Chrome options for stability
            options = uc.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--no-sandbox")
            
            # Create driver with auto-update (undetected-chromedriver handles version matching)
            try:
                self.driver = uc.Chrome(options=options, version_main=None)  # Auto-detect Chrome version
            except Exception as e:
                # Fallback: Force update
                messagebox.showwarning(
                    "Chrome Driver Update",
                    "Updating Chrome driver to match your Chrome version...\nThis may take a moment.",
                    parent=self
                )
                self.driver = uc.Chrome(options=options, version_main=None, driver_executable_path=None)
            
            self.driver.maximize_window()
            self.driver.get("https://web.whatsapp.com")
            
            # Wait for user login
            messagebox.showinfo(
                "WhatsApp Login Required", 
                "Please scan the QR code in Chrome to login.\n\n"
                "Click OK ONLY after chats are loaded.",
                parent=self
            )
            
            wait = self.WebDriverWait(self.driver, 120)
            
            # Multiple selectors for chat panel (future-proof)
            chat_panel_selectors = [
                "//div[@id='pane-side']",
                "//div[@id='side']",
                "//div[@role='grid']",
                "//div[contains(@class, 'chat-list')]"
            ]
            
            login_success = False
            for selector in chat_panel_selectors:
                try:
                    wait.until(self.EC.presence_of_element_located((self.By.XPATH, selector)))
                    login_success = True
                    break
                except:
                    continue
            
            if not login_success:
                raise Exception("Could not detect WhatsApp login")
            
            time.sleep(2)
            
            messagebox.showinfo(
                "Login Successful", 
                "WhatsApp is ready! Messages will now be sent.",
                parent=self
            )
            return True
            
        except Exception as e:
            self._ui_safe(lambda: messagebox.showerror(
            "WhatsApp Login Failed",
            "Could not detect WhatsApp Web login.\n\n"
            "Please make sure:\n"
            "• Internet is connected\n"
            "• WhatsApp Web QR is scanned\n"
            "• Chrome window is not closed\n\n"
            "Then try again.",
            parent=self
            ))
            return False


            try:
                if self.driver:
                    self.driver.quit()
            except:
                pass
            self.driver = None
            return False
    
    def _ui_safe(self, func):
        """Safely execute UI code from worker threads"""
        try:
            if self.winfo_exists():
                self.after(0, func)
        except:
            pass











    # message_tab.py  — Part 3/3 (Updated)

    def send_messages(self):
        if self.is_sending:
            messagebox.showwarning("Already Sending", "Messages are currently being sent. Please wait.", parent=self)
            return
        if not self.check_internet():
            messagebox.showerror("No Internet", "Check your Internet", parent=self)
            return

        selected_customers = [cb['data'] for cb in self.checkboxes.values() if cb['var'].get()]
        if not selected_customers:
            messagebox.showwarning("No Selection", "Please select at least one customer.", parent=self)
            return

        confirm = messagebox.askyesno("Confirm Send",
                                    f"📱 Send WhatsApp messages to {len(selected_customers)} customers?\n\n"
                                    f"⏱️ Estimated time: {max(1, (len(selected_customers) * 9) // 60)} minutes\n\n"
                                    "⚠️ Do NOT close the Chrome window during sending", parent=self)
        if not confirm:
            return

        # Change button to "Initializing..."
        try:
            self.btn_send.configure(state="disabled", text="⏳ Initializing...")
        except:
            pass

        self.is_sending = True
        self.is_paused = False
        self.pause_requested = False
        self.failed_list = []
        
        # Load existing sent_list
        if not hasattr(self, 'sent_list') or self.sent_list is None:
            self.sent_list = []
        self.load_sent_list()

        # Store count BEFORE starting (for accurate reporting)
        self.sent_count_before = len(self.sent_list)

        self.send_thread = threading.Thread(target=self.send_messages_thread, args=(selected_customers, False), daemon=False)
        self.send_thread.start()

    def stop_background_tasks(self):
        self.is_sending = False
        self.pause_requested = True

    def send_messages_thread(self, selected_customers, is_retry=False):
        try:
            # Initialize Selenium and wait for user to login
            if not self.init_selenium():
                self.after(0, lambda: messagebox.showerror(
                    "Initialization Failed",
                    "Could not initialize WhatsApp. Sending cancelled.",
                    parent=self
                ))
                self._ui_safe(lambda: self.reset_send_button(is_retry))

                return

            self.is_sending = True
            self.pause_requested = False
            self.failed_list = []  # Initialize failed list

            total = len(selected_customers)
            self.after(0, lambda: self.create_progress_window(total))
            time.sleep(0.5)

            # Load existing sent list and track starting count
            self.load_sent_list()
            sent_count_before = len(self.sent_list)

            # Track successful sends in this session
            session_sent_count = 0
            session_failed_count = 0

            for idx, cust in enumerate(selected_customers):

                if not self.is_sending or self.pause_requested:
                    break

                while getattr(self, "is_paused", False):
                    time.sleep(0.25)

                self.after(
                    0,
                    lambda i=idx+1, t=total, c=cust:
                        self.update_progress_safe(i, t, c.get("Name",""))
                )

                phone = self.format_phone_number(cust.get("Phone"))
                if not phone:
                    self.failed_list.append({**cust, "Reason": "Invalid number"})
                    session_failed_count += 1
                    continue


                msg = self.create_message(cust)
                pyperclip.copy(msg)
                time.sleep(0.3)

                try:
                    self.send_whatsapp_message(phone, msg)
                    # Successfully sent - add to sent_list
                    self.append_to_sent_list(cust)
                    session_sent_count += 1
                    time.sleep(5)  # Delay only after successful send
                except Exception as e:
                    error_msg = str(e)
                    self.failed_list.append({**cust, "Reason": error_msg})
                    session_failed_count += 1

                    # Quick skip for invalid numbers (no long delay)
                    if "invalid" in error_msg.lower() or "not load" in error_msg.lower():
                        time.sleep(0.5)  # Minimal delay for invalid numbers
                    else:
                        time.sleep(1)  # Short delay for other errors
                    continue

            # Save both lists
            self.save_sent_list()
            self.save_unsent_list()
            
            # Pass session counts to finalize (not cumulative totals)
            self.after(0, lambda: self.finalize_sending(session_sent_count, session_failed_count))

        except Exception as e:
            self._ui_safe(lambda e=str(e): messagebox.showerror(
            "Sending Error",
            f"An error occurred while sending messages:\n\n{e}",
            parent=self
            ))

            # Still show what was processed before error
            if hasattr(self, 'failed_list'):
                session_failed = len(self.failed_list)
            else:
                session_failed = 0
            self.after(0, lambda: self.finalize_sending(session_sent_count if 'session_sent_count' in locals() else 0, session_failed))
        finally:
            self.is_sending = False
            self.send_thread = None
            self._ui_safe(lambda: self.reset_send_button(is_retry))



    def send_whatsapp_message(self, phone, message):
        
        if not self.driver:
            raise Exception("Driver not initialized")

        try:
            # ========== STEP 1: NAVIGATE ==========
            self.driver.get(f"https://web.whatsapp.com/send?phone={phone}")
            time.sleep(2)  # Reduced from 4
            
            # ========== STEP 2: QUICK INVALID CHECK (3 seconds max) ==========
            check_start = time.time()
            while time.time() - check_start < 3:
                try:
                    page_text = self.driver.find_element(self.By.TAG_NAME, "body").text.lower()
                    if "invalid" in page_text or "not on whatsapp" in page_text:
                        raise Exception("Invalid phone number")
                    
                    # If footer exists, break immediately
                    if self.driver.find_elements(self.By.XPATH, "//footer"):
                        break
                        
                    time.sleep(0.2)
                except:
                    break  # Don't waste time on errors
            
            # ========== STEP 3: FIND INPUT BOX (10 seconds max) ==========
            input_box = None
            wait_start = time.time()
            
            while time.time() - wait_start < 10:
                try:
                    # Try multiple selectors quickly
                    for selector in [
                        "//div[@contenteditable='true'][@data-tab='10']",
                        "//footer//div[@contenteditable='true']",
                        "//div[@role='textbox']"
                    ]:
                        elements = self.driver.find_elements(self.By.XPATH, selector)
                        if elements and elements[0].is_displayed():
                            input_box = elements[0]
                            break
                    
                    if input_box:
                        break
                    
                    time.sleep(0.3)
                except:
                    time.sleep(0.3)
            
            if not input_box:
                raise Exception("Input not found")
            
            # ========== STEP 4: PASTE & SEND (Fast) ==========
            input_box.click()
            time.sleep(0.5)
            
            # Paste
            pyperclip.copy(message)
            time.sleep(0.2)
            input_box.send_keys(self.Keys.CONTROL, "v")
            time.sleep(0.8)
            
            # Send immediately
            try:
                # Try send button first
                send_btn = self.driver.find_element(self.By.XPATH, "//button[@aria-label='Send']")
                send_btn.click()
            except:
                # Fallback: Enter key
                input_box.send_keys(self.Keys.ENTER)
            
            time.sleep(1)
            
            # ========== STEP 5: QUICK VERIFICATION (5 seconds max) ==========
            verify_start = time.time()
            
            while time.time() - verify_start < 5:
                try:
                    # Just check if ANY tick exists
                    ticks = self.driver.find_elements(self.By.XPATH, "//span[@data-icon='msg-check']")
                    if ticks:
                        return True  # Success!
                    
                    # Or check if input cleared
                    if len(input_box.text) < 10:
                        return True  # Likely sent
                    
                    time.sleep(0.3)
                except:
                    break
            
            # If 5 seconds passed without error, assume success
            return True
                
        except Exception as e:
            error_msg = str(e).lower()
            
            # Invalid number - fail fast
            if "invalid" in error_msg or "not on whatsapp" in error_msg:
                raise Exception("Invalid phone number")
            
            # Other errors
            raise Exception(f"Failed: {e}")
    
    def show_range_selector(self):
        """Show range selection dialog with visible buttons"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Select Range")
        dialog.geometry("420x280")
        dialog.transient(self)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 210
        y = (dialog.winfo_screenheight() // 2) - 140
        dialog.geometry(f"420x280+{x}+{y}")
        
        # Root container
        root = ctk.CTkFrame(dialog, fg_color="#f8fafc")
        root.pack(fill="both", expand=True)
        
        # Content area
        content = ctk.CTkFrame(root, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=30, pady=20)
        
        ctk.CTkLabel(
            content,
            text="🎯 Select Customer Range",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=self.colors["primary"]
        ).pack(pady=(0, 20))
        
        # Range inputs
        range_frame = ctk.CTkFrame(content, fg_color="transparent")
        range_frame.pack(pady=15)
        
        ctk.CTkLabel(
            range_frame,
            text="From:",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(side="left", padx=5)
        
        ent_from = ctk.CTkEntry(
            range_frame,
            width=80,
            height=45,
            font=ctk.CTkFont(size=16),
            justify="center"
        )
        ent_from.insert(0, "1")
        ent_from.pack(side="left", padx=5)
        
        ctk.CTkLabel(
            range_frame,
            text=" — ",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(side="left")
        
        ctk.CTkLabel(
            range_frame,
            text="To:",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(side="left", padx=5)
        
        ent_to = ctk.CTkEntry(
            range_frame,
            width=80,
            height=45,
            font=ctk.CTkFont(size=16),
            justify="center"
        )
        ent_to.insert(0, str(min(50, len(self.customer_data))))
        ent_to.pack(side="left", padx=5)
        
        # Max label
        ctk.CTkLabel(
            content,
            text=f"(Maximum: {len(self.customer_data)} customers)",
            font=ctk.CTkFont(size=13),
            text_color="#6b7280"
        ).pack(pady=8)
        
        def apply_range():
            try:
                from_val = int(ent_from.get().strip())
                to_val = int(ent_to.get().strip())
                
                # Validation
                if from_val < 1:
                    messagebox.showerror("Invalid Range", "Start must be at least 1", parent=dialog)
                    return
                
                if to_val > len(self.customer_data):
                    messagebox.showerror("Invalid Range", f"End cannot exceed {len(self.customer_data)}", parent=dialog)
                    return
                
                if from_val > to_val:
                    messagebox.showerror("Invalid Range", "Start cannot be greater than End", parent=dialog)
                    return
                
                # Clear all selections
                for cb in self.checkboxes.values():
                    cb['var'].set(False)
                self.previously_selected.clear()
                
                # Select range (S.No is 1-indexed)
                for iid, cid_key in self.tree_iid_to_key.items():
                    vals = self.tree.item(iid).get('values', [])
                    if vals:
                        sno = int(vals[1])  # S.No column
                        if from_val <= sno <= to_val:
                            if cid_key in self.checkboxes:
                                self.checkboxes[cid_key]['var'].set(True)
                                self.previously_selected.add(cid_key)
                
                self.update_checkbox_display()
                self.update_selected_count()
                
                dialog.destroy()
                messagebox.showinfo(
                    "Range Selected",
                    f"✓ Selected customers {from_val} to {to_val}\n\nTotal: {to_val - from_val + 1} customers",
                    parent=self
                )
                
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid numbers", parent=dialog)
        
        # ✅ FIXED BUTTON CONTAINER
        button_container = ctk.CTkFrame(root, fg_color="transparent")
        button_container.pack(fill="x", padx=30, pady=(0, 20))
        
        ctk.CTkButton(
            button_container,
            text="✓ Apply Range",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=150,
            height=50,
            corner_radius=14,
            fg_color="#10b981",
            hover_color="#059669",
            command=apply_range
            ).pack(side="left", padx=10)
            
        ctk.CTkButton(
                button_container,
                text="✕ Cancel",
                font=ctk.CTkFont(size=16, weight="bold"),
                width=150,
                height=50,
                corner_radius=14,
                fg_color="#6b7280",
                hover_color="#4b5563",
                command=dialog.destroy
        ).pack(side="left", padx=10)









    # ---------------------------
    # Progress & finalization
    # ---------------------------
    def create_progress_window(self, total):
        # destroy previous window if exists
        try:
            if hasattr(self, "progress_window") and self.progress_window.winfo_exists():
                self.progress_window.destroy()
        except:
            pass

        self.progress_window = ctk.CTkToplevel(self)
        self.progress_window.title("Sending Messages")
        self.progress_window.geometry("600x300")
        self.progress_window.transient(self)
        self.progress_window.grab_set()

        # ----- CLOSE BUTTON (X) instantly stops thread -----
        def on_close():
            self.pause_requested = True
            self.is_sending = False
            try:
                self.progress_window.destroy()
            except:
                pass

            # Immediately finalize
            self.finalize_sending()

        self.progress_window.protocol("WM_DELETE_WINDOW", on_close)

        # Labels & progress bar
        title = ctk.CTkLabel(self.progress_window, text="Sending Messages...", font=("Arial", 18, "bold"))
        title.pack(pady=10)

        self.progress_label = ctk.CTkLabel(self.progress_window, text="Preparing...", font=("Arial", 14))
        self.progress_label.pack(pady=5)

        self.progress_bar = ctk.CTkProgressBar(self.progress_window, width=500)
        self.progress_bar.pack(pady=10)
        self.progress_bar.set(0)

        # -------- PAUSE / RESUME BUTTON --------
        self.pause_btn = ctk.CTkButton(self.progress_window, text="⏸ Pause", command=self.toggle_pause_resume)
        self.pause_btn.pack(pady=10)



    def update_progress_safe(self, current, total, customer_name):
        if hasattr(self, 'progress_window') and self.progress_window.winfo_exists():
            try:
                self.progress_bar.set(current / max(1, total))
                self.progress_label.configure(text=f"Sending to: {customer_name}\n({current}/{total})")
                self.progress_window.update()
            except:
                pass

    def toggle_pause_resume(self):
        """Pause or resume sending."""
        if not hasattr(self, "is_paused"):
            self.is_paused = False

        if self.is_paused:
            # RESUME
            self.is_paused = False
            self.pause_requested = False
            try:
                self.pause_btn.configure(text="⏸ Pause")
            except:
                pass
        else:
            # PAUSE
            self.is_paused = True
            self.pause_requested = True
            try:
                self.pause_btn.configure(text="▶ Resume")
            except:
                pass


    def finalize_sending(self, sent_count=None, failed_count=None):
        """Finalize sending and show results."""
        try:
            if hasattr(self, "progress_window") and self.progress_window.winfo_exists():
                self.progress_window.destroy()
        except:
            pass

        # Calculate counts if not provided
        if sent_count is None:
            # Count customers that were actually sent in this session
            if hasattr(self, 'sent_list') and self.sent_list:
                sent_count = len(self.sent_list)
            else:
                sent_count = 0
        
        if failed_count is None:
            # Count customers that failed in this session
            if hasattr(self, 'failed_list') and self.failed_list:
                failed_count = len(self.failed_list)
            else:
                failed_count = 0

        # Save both lists
        try:
            self.save_sent_list()
            self.save_unsent_list()
        except Exception as e:
            print(f"Error saving lists: {e}")

        # Calculate total processed
        total_processed = sent_count + failed_count

        # Show summary with counts from THIS sending session only
        try:
            if total_processed > 0:
                messagebox.showinfo(
                    "Sending Complete",
                    f"✅ Successfully Sent: {sent_count}\n"
                    f"❌ Failed: {failed_count}\n\n"
                    f"Total Processed: {total_processed} customers",
                    parent=self
                )
            else:
                messagebox.showinfo(
                    "Sending Complete",
                    "No messages were sent in this session.",
                    parent=self
                )
        except:
            pass

        self.is_sending = False
        self.pause_requested = False
        self.send_thread = None

    def reset_send_button(self, is_retry=False):
        self.is_sending = False
        self.is_paused = False
        self.pause_requested = False
        try:
            if is_retry and hasattr(self, 'unsent_btn_send'):
                if self.unsent_btn_send.winfo_exists():
                    self.unsent_btn_send.configure(state="normal", text="📱 Retry Failed Messages")
            elif hasattr(self, 'btn_send'):
                if self.btn_send.winfo_exists():
                    self.btn_send.configure(state="normal", text="📱 Send to WhatsApp")
        except:
            pass

    # ---------------------------
    # Message create + saving
    # ---------------------------
    def create_message(self, customer):
        """Create personalized message using latest totals from customer dict."""
        days_in_month = calendar.monthrange(self.selected_year, self.selected_month)[1] if self.selected_month else 30
        month_name = self.month_names[self.selected_month-1] if self.selected_month else "Month"

        name = customer.get('Name') or customer.get('name') or "Customer"
        cid = customer.get('CID') or customer.get('cid') or ""
        
        # Get total liters
        try:
            total_ltr = float(customer.get('Total_Ltr', customer.get('total_ltr', 0)) or 0)
        except:
            total_ltr = 0.0
        
        # Get total amount
        try:
            total_amt = customer.get('Total_Amt') or customer.get('total_amt') or customer.get('Total_RT') or 0
            total_amt = float(total_amt) if total_amt else 0.0
        except:
            total_amt = 0.0

        # UPDATED: Use config values
        lines = [
            f"🥛 {self.business_name} - Monthly Bill",
            "",
            f"Dear {name},",
            "",
            f"🆔 Customer ID: {cid}",
            f"📅 Period: 01 {month_name[:3].upper()} {self.selected_year} - {days_in_month} {month_name[:3].upper()} {self.selected_year}",
            "",
            "📊 Delivery Summary:",
            "━━━━━━━━━━━━━━━━━━",
            f"🥛 Total Milk: {total_ltr:.2f} Liters",
            f"💰 Amount Due: ₹{total_amt:.2f}",
            "━━━━━━━━━━━━━━━━━━",
            "",
            f"📞 For queries: {self.contact_number}",
            f"💳 Pay via: {self.payment_info}",
            "",
            "Thank you! 🙏"
        ]
        return "\n".join(lines)

    def get_unsent_file_path(self):
        """Return the path for this month's unsent file in Status/{Month}_{Year}/ folder."""
        if not self.selected_month or not self.selected_year:
            return None
        month_name = self.month_names[self.selected_month-1].title()
        month_folder = os.path.join(self.status_folder, f"{month_name}_{self.selected_year}")
        os.makedirs(month_folder, exist_ok=True)
        return os.path.join(month_folder, f"Unsent_{month_name}_{self.selected_year}.xlsx")
    
    def append_to_sent_list(self, customer):
        """Add customer to sent_list with proper data structure."""
        try:
            # Check if already in sent_list
            cid = customer.get('CID', customer.get('cid', ''))
            if any(c.get('CID') == cid for c in self.sent_list):
                return
            
            # Get total amount from all possible keys
            total_amt = customer.get('Total_Amt') or customer.get('total_amt') or customer.get('Total_RT') or 0
            try:
                total_amt = float(total_amt) if total_amt else 0.0
            except:
                total_amt = 0.0
            
            # Get total liters
            total_ltr = customer.get('Total_Ltr', customer.get('total_ltr', 0)) or 0
            try:
                total_ltr = float(total_ltr) if total_ltr else 0.0
            except:
                total_ltr = 0.0
            
            # Add to sent_list with all required fields
            self.sent_list.append({
                "CID": cid,
                "Name": customer.get('Name', customer.get('name', '')),
                "Phone": customer.get('Phone', customer.get('phone', '')),
                "Total_Ltr": total_ltr,
                "Total_Amt": total_amt
            })
        except Exception as e:
            print(f"Error appending to sent_list: {e}")
    
    def save_unsent_list(self):
        """
        Calculate unsent as: all customers from monthly sheet - sent_list.
        Also includes failed_list from current session.
        Save to unsent file in month-specific folder.
        """
        try:
            unsent_file = self.get_unsent_file_path()
            if not unsent_file:
                return
            
            # Get all CIDs that have been successfully sent
            sent_cids = set([str(c.get('CID') or "").strip() for c in self.sent_list if c.get('CID')])
            
            # Calculate unsent from ALL customers in monthly sheet
            unsent_data = []
            for customer in self.customer_data:
                cid = str(customer.get('CID') or "").strip()
                if cid and cid not in sent_cids:
                    # Get total amount properly
                    total_amt = customer.get('Total_Amt') or customer.get('total_amt') or customer.get('Total_RT') or 0
                    try:
                        total_amt = float(total_amt) if total_amt else 0.0
                    except:
                        total_amt = 0.0
                    
                    unsent_data.append({
                    "CID": cid,
                    "Name": customer.get("Name", ""),
                    "Phone": customer.get("Phone", ""),
                    "Total_Ltr": float(customer.get("Total_Ltr", 0) or 0),
                    "Total_Amt": total_amt,
                    "Reason": "Not selected"
                    })

            
            # Add current session's failed customers with their reasons
            if hasattr(self, 'failed_list') and self.failed_list:
                for failed in self.failed_list:
                    cid = str(failed.get('CID') or "").strip()
                    # Only add if not already in unsent_data and not sent
                    if cid and cid not in sent_cids:
                        already_added = any(u['CID'] == cid for u in unsent_data)
                        if not already_added:
                            # Get total amount properly
                            total_amt = failed.get('Total_Amt') or failed.get('total_amt') or failed.get('Total_RT') or 0
                            try:
                                total_amt = float(total_amt) if total_amt else 0.0
                            except:
                                total_amt = 0.0
                            
                            unsent_data.append({
                                "CID": cid,
                                "Name": failed.get("Name", ""),
                                "Phone": failed.get("Phone", ""),
                                "Total_Ltr": float(failed.get("Total_Ltr", 0) or 0),
                                "Total_Amt": total_amt,
                                "Reason": failed.get("Reason", "")
                            })
                        else:
                            # Update reason if already exists
                            for u in unsent_data:
                                if u['CID'] == cid:
                                    u['Reason'] = failed.get("Reason", "")
                                    break
            
            if unsent_data:
                df = pd.DataFrame(unsent_data)
                df.to_excel(unsent_file, index=False, engine="openpyxl")
            else:
                # If all sent, delete unsent file if exists
                if os.path.exists(unsent_file):
                    os.remove(unsent_file)
        except Exception as e:
            print(f"Failed to save unsent list: {e}")

    # ---------------------------
    # Refresh
    # ---------------------------
    def refresh_data(self):
        if not self.selected_month:
            messagebox.showwarning("No Month Selected", "Please select a month to refresh.", parent=self)
            return
        try:
            # reload sheet into memory & UI
            self.load_customer_list(self.selected_month)
            messagebox.showinfo("Refreshed", "Customer data refreshed for the selected month.", parent=self)
        except Exception as e:
            messagebox.showerror("Refresh Failed", f"Failed to refresh data: {e}", parent=self)